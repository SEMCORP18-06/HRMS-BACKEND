import os
import csv
import datetime
import io
from flask import Flask, request, jsonify, g, send_from_directory, send_file
from flask_cors import CORS
import jwt
from bson.objectid import ObjectId

from .database import db, seed_db, serialize_doc, serialize_list
from .utils.pdf_encrypt import generate_salary_breakup_pdf, encrypt_pdf_aes
from .utils.mailer import send_email
from .utils.scheduler import init_scheduler, queue_event_reminders, check_and_send_celebrations
from .utils.ctc_export import generate_pdf, generate_excel, generate_word

# Initialize and seed database using MongoDB Atlas
seed_db()
init_scheduler()

# Configure Flask to serve the Vite frontend build folder
# Fallback to local static folder if frontend/dist doesn't exist (e.g. on Vercel serverless deployment)
_local_static = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "static"))
if os.path.exists(os.path.join(_local_static, "index.html")):
    frontend_dist = _local_static
else:
    frontend_dist = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "frontend", "dist"))

# NOTE: static_folder=None so Flask does NOT register its own static handler.
# All static file serving (assets, favicon, etc.) is done in the catch_all route below,
# which guarantees SPA routes like /invite/activate always return index.html.
app = Flask(__name__, static_folder=None)
CORS(app)

def get_upload_dir(*paths):
    if os.environ.get("VERCEL") or os.environ.get("VERCEL_ENV"):
        upload_dir = os.path.join("/tmp", "static", "uploads", *paths)
    else:
        upload_dir = os.path.join(app.root_path, "static", "uploads", *paths)
    os.makedirs(upload_dir, exist_ok=True)
    return upload_dir
import base64

def convert_file_to_base64_uri(file, allowed_ext=None):
    if not file or file.filename == '':
        return None
    fname = file.filename.lower()
    ext = os.path.splitext(fname)[1]
    if allowed_ext and ext not in allowed_ext:
        return None
    file_bytes = file.read()
    file.seek(0)
    
    mime_type = "application/octet-stream"
    if ext == ".pdf":
        mime_type = "application/pdf"
    elif ext in [".jpg", ".jpeg"]:
        mime_type = "image/jpeg"
    elif ext == ".png":
        mime_type = "image/png"
    elif ext == ".webp":
        mime_type = "image/webp"
    elif ext == ".gif":
        mime_type = "image/gif"
    elif ext == ".svg":
        mime_type = "image/svg+xml"
    elif ext == ".docx":
        mime_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif ext == ".xlsx":
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
    base64_data = base64.b64encode(file_bytes).decode('utf-8')
    return f"data:{mime_type};base64,{base64_data}"

@app.before_request
def handle_options_preflight():
    if request.method == "OPTIONS":
        response = app.make_response("")
        response.headers["Access-Control-Allow-Origin"] = "*"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
        return response

@app.after_request
def handle_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    if "Access-Control-Allow-Headers" not in response.headers:
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    if "Access-Control-Allow-Methods" not in response.headers:
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    return response

SECRET_KEY = "hr-ops-secret-key-12345"

# --- JWT Helpers ---
def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.datetime.utcnow() + datetime.timedelta(days=1)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm="HS256")

def decode_access_token(token: str):
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
    except jwt.PyJWTError:
        return None

# --- Decorator for routes ---
from functools import wraps
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = None
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            if auth_header.startswith('Bearer '):
                token = auth_header.split(' ')[1]
        elif 'token' in request.args:
            token = request.args.get('token')
        
        if not token:
            return jsonify({"detail": "Token is missing"}), 401
            
        payload = decode_access_token(token)
        if not payload:
            return jsonify({"detail": "Invalid or expired token"}), 401
            
        g.current_user = payload
        return f(*args, **kwargs)
    return decorated_function

# --- API Routes ---

@app.route('/api/auth/tenants', methods=['GET'])
def get_tenants():
    try:
        tenants = list(db.tenants.find({}))
        # Map _id (since it's a string in our seeder, e.g., 'acme')
        result = []
        for t in tenants:
            result.append({
                "id": t["_id"],
                "name": t["name"],
                "sso_domain": t["sso_domain"]
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/auth/login', methods=['POST'])
def sso_login():
    data = request.json or {}
    email = data.get('email')
    password = data.get('password')
    role_pref = data.get('role')
    
    if not email or not password:
        return jsonify({"detail": "Email and password are required"}), 400
        
    try:
        email_domain = email.split('@')[-1] if '@' in email else ''
        is_personal_allowed = False
        if email_domain.lower() != "semcogroups.com":
            allowed_emp = db.employees.find_one({
                "$or": [{"email": email}, {"personal_email": email}],
                "allow_personal_email_access": True,
                "tenant_id": "semco"
            })
            if allowed_emp:
                is_personal_allowed = True
            if not is_personal_allowed:
                return jsonify({"detail": "Only semcogroups.com domain accounts are permitted."}), 400
            
        employees = list(db.employees.find({
            "$or": [{"email": email}, {"personal_email": email}],
            "tenant_id": "semco"
        }))
        if not employees:
            return jsonify({"detail": "Invalid email or password."}), 401
            
        employee = None
        # Match using preferred role if specified
        if role_pref:
            for emp in employees:
                if emp.get("password") == password and emp.get("role") == role_pref:
                    employee = emp
                    break
                    
        # Fallback if no specific role matched or preferred
        if not employee:
            for emp in employees:
                if emp.get("password") == password and emp.get("role") == "Admin (HR)":
                    employee = emp
                    break
        if not employee:
            for emp in employees:
                if emp.get("password") == password:
                    employee = emp
                    break

        if not employee:
            return jsonify({"detail": "Invalid email or password."}), 401
            
        token = create_access_token({
            "employee_id": str(employee["_id"]),
            "name": employee["name"],
            "email": employee.get("email") or employee.get("personal_email"),
            "role": employee["role"],
            "tenant_id": "semco"
        })
        
        return jsonify({
            "access_token": token,
            "user": {
                "id": str(employee["_id"]),
                "name": employee.get("name", ""),
                "email": employee.get("email") or employee.get("personal_email"),
                "role": employee.get("role", "Employee"),
                "salutation": employee.get("salutation", ""),
                "emp_id": employee.get("emp_id", ""),
                "department": employee.get("department", ""),
                "designation": employee.get("designation", ""),
                "doj": employee.get("joining_date") or employee.get("doj", ""),
                "dob": employee.get("birthday") or employee.get("dob", ""),
                "age": employee.get("age", ""),
                "personal_email": employee.get("personal_email", ""),
                "current_address": employee.get("current_address", ""),
                "office_contact": employee.get("office_contact", ""),
                "personal_contact": employee.get("personal_contact", "")
            },
            "tenant": {"id": "semco", "name": "SEMCO Groups"}
        })
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/auth/signup', methods=['POST'])
def sso_signup():
    data = request.json or {}
    name = data.get('name')
    email = data.get('email')
    password = data.get('password')
    role = data.get('role')
    department = data.get('department', 'General')
    birthday = data.get('birthday')
    anniversary = data.get('anniversary')
    
    if not name or not email or not password:
        return jsonify({"detail": "Name, email, and password are required"}), 400
        
    try:
        email_domain = email.split('@')[-1] if '@' in email else ''
        is_personal_allowed = False
        if email_domain.lower() != "semcogroups.com":
            allowed_emp = db.employees.find_one({
                "personal_email": email,
                "allow_personal_email_access": True,
                "tenant_id": "semco"
            })
            if allowed_emp:
                is_personal_allowed = True
            if not is_personal_allowed:
                return jsonify({"detail": "Only semcogroups.com domain accounts are permitted."}), 400
            
        admin_exists = db.employees.count_documents({"role": "Admin (HR)"}) > 0
        
        if admin_exists:
            # Look for an unactivated profile (which has no password set yet) matching the email
            existing = db.employees.find_one({
                "$or": [{"email": email}, {"personal_email": email}],
                "$or": [
                    {"password": {"$exists": False}},
                    {"password": None},
                    {"password": ""}
                ]
            })
            if not existing:
                # If no unactivated profile exists, check if it's already registered
                already_registered = db.employees.find_one({
                    "$or": [{"email": email}, {"personal_email": email}],
                    "password": {"$ne": ""}
                })
                if already_registered:
                    return jsonify({"detail": "Account with this email already exists."}), 400
                return jsonify({"detail": "Your company email has not been provisioned by HR yet. Please contact your HR administrator."}), 400

            # Only update password; preserve all onboarding fields (name, role, dob, doj, department, etc.)
            update_fields = {"password": password, "status": "ACTIVE"}
            if not existing.get("email") or existing.get("email") == "":
                update_fields["email"] = email
            # Only fill name if it was blank in the existing record
            if not existing.get("name"):
                update_fields["name"] = name
            db.employees.update_one(
                {"_id": existing["_id"]},
                {"$set": update_fields}
            )
            employee = db.employees.find_one({"_id": existing["_id"]})
        else:
            existing = db.employees.find_one({"email": email})
            if existing:
                return jsonify({"detail": "Account with this email already exists."}), 400
                
            if not role or role not in ["Admin (HR)", "Employee"]:
                role = "Admin (HR)"
            if role == "Admin (HR)":
                department = "HR Ops"
                
            today = datetime.date.today().isoformat()
            res = db.employees.insert_one({
                "tenant_id": "semco",
                "name": name,
                "email": email,
                "password": password,
                "role": role,
                "department": department,
                "birthday": birthday or None,
                "anniversary": anniversary or None,
                "joining_date": today,
                "status": "ACTIVE",
                "system_access_revoked": 0,
                "allow_personal_email_access": False
            })
            employee = db.employees.find_one({"_id": res.inserted_id})
            
        token = create_access_token({
            "employee_id": str(employee["_id"]),
            "name": employee["name"],
            "email": employee["email"],
            "role": employee.get("role", "Employee"),
            "tenant_id": "semco"
        })
        
        return jsonify({
            "access_token": token,
            "user": {
                "id": str(employee["_id"]),
                "name": employee.get("name", ""),
                "email": employee.get("email", ""),
                "role": employee.get("role", "Employee"),
                "salutation": employee.get("salutation", ""),
                "emp_id": employee.get("emp_id", ""),
                "department": employee.get("department", ""),
                "designation": employee.get("designation", ""),
                "doj": employee.get("joining_date") or employee.get("doj", ""),
                "dob": employee.get("birthday") or employee.get("dob", ""),
                "age": employee.get("age", ""),
                "personal_email": employee.get("personal_email", ""),
                "current_address": employee.get("current_address", ""),
                "office_contact": employee.get("office_contact", ""),
                "personal_contact": employee.get("personal_contact", "")
            },
            "tenant": {"id": "semco", "name": "SEMCO Groups"}
        })
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/auth/me', methods=['GET'])
@login_required
def auth_me():
    try:
        emp_id = g.current_user["employee_id"]
        emp = db.employees.find_one({"_id": ObjectId(emp_id)})
        
        tenant_id = g.current_user["tenant_id"]
        tenant = db.tenants.find_one({"_id": tenant_id})
        
        if not emp or not tenant:
            return jsonify({"detail": "User not found"}), 404
            
        return jsonify({
            "user": {
                "id": str(emp["_id"]),
                "salutation": emp.get("salutation", ""),
                "name": emp.get("name", ""),
                "emp_id": emp.get("emp_id", ""),
                "email": emp.get("email", ""),
                "role": emp.get("role", ""),
                "department": emp.get("department", ""),
                "designation": emp.get("designation", ""),
                "doj": emp.get("joining_date", "") or emp.get("doj", ""),
                "dob": emp.get("birthday", "") or emp.get("dob", ""),
                "age": emp.get("age", ""),
                "personal_email": emp.get("personal_email", ""),
                "current_address": emp.get("current_address", ""),
                "office_contact": emp.get("office_contact", ""),
                "personal_contact": emp.get("personal_contact", ""),
                "allow_late_attendance_marking": emp.get("allow_late_attendance_marking", False)
            },
            "tenant": {"id": tenant["_id"], "name": tenant["name"]}
        })
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/employees/me', methods=['PUT'])
@login_required
def update_my_profile():
    try:
        emp_id = g.current_user["employee_id"]
        data = request.json or {}
        
        salutation = data.get('salutation', '')
        name = data.get('name', '')
        emp_code = data.get('emp_code') or data.get('emp_id', '')
        department = data.get('department', '')
        designation = data.get('designation', '')
        doj = data.get('doj', '')
        dob = data.get('dob', '')
        age = data.get('age', '')
        personal_email = data.get('personal_email', '')
        current_address = data.get('current_address', '')
        office_contact = data.get('office_contact', '')
        personal_contact = data.get('personal_contact', '')
        
        if dob and not age:
            try:
                dob_date = datetime.date.fromisoformat(str(dob)[:10])
                today = datetime.date.today()
                age = today.year - dob_date.year - ((today.month, today.day) < (dob_date.month, dob_date.day))
            except Exception:
                pass
                
        update_set = {
            "salutation": salutation,
            "name": name,
            "emp_id": emp_code,
            "department": department,
            "designation": designation,
            "joining_date": doj,
            "doj": doj,
            "birthday": dob,
            "dob": dob,
            "age": age,
            "personal_email": personal_email,
            "current_address": current_address,
            "office_contact": office_contact,
            "personal_contact": personal_contact
        }
        
        db.employees.update_one(
            {"_id": ObjectId(emp_id)},
            {"$set": update_set}
        )
        return jsonify({"message": "My profile updated successfully."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

def parse_date_dm4y(date_val):
    if not date_val:
        return None
    if isinstance(date_val, (datetime.date, datetime.datetime)):
        return date_val.isoformat()[:10]
    try:
        if isinstance(date_val, (int, float)) and date_val > 1000:
            from openpyxl.utils.datetime import from_excel
            return from_excel(date_val).date().isoformat()
    except Exception:
        pass
    date_str = str(date_val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.datetime.strptime(date_str, fmt).date().isoformat()
        except ValueError:
            pass
    if len(date_str) >= 10 and date_str[4] == '-' and date_str[7] == '-':
        return date_str[:10]
    return None

def normalize_keys(row_dict):
    normalized = {}
    for k, v in row_dict.items():
        if k is None:
            continue
        k_str = str(k).strip().lower()
        if k_str in ('emp. id', 'emp id', 'employee id', 'id', 'sr. no', 'sr no'):
            normalized['emp_id'] = v
        elif k_str in ('name', 'full name', 'employee name', 'emp name'):
            normalized['name'] = v
        elif k_str in ('email', 'email address', 'company mail id', 'company email', 'company mail', 'work email'):
            normalized['email'] = v
        elif k_str in ('password',):
            normalized['password'] = v
        elif k_str in ('designation', 'role', 'job title', 'title', 'post'):
            normalized['designation'] = v
            normalized['role'] = v
        elif k_str in ('department', 'dept'):
            normalized['department'] = v
        elif k_str in ('dob', 'birthday', 'birth date', 'date of birth'):
            normalized['dob'] = v
        elif k_str in ('doj', 'joining date', 'date of joining', 'anniversary'):
            normalized['doj'] = v
        elif k_str in ('personal mail id', 'personal email', 'personal mail', 'mail id', 'email id', 'mail', 'personal email id'):
            normalized['personal_email'] = v
        elif k_str in ('current address', 'address', 'residence'):
            normalized['current_address'] = v
        elif k_str in ('office contact', 'office phone', 'work contact', 'work phone'):
            normalized['office_contact'] = v
        elif k_str in ('personal contact', 'personal phone', 'contact', 'mobile', 'mobile number', 'phone', 'phone number'):
            normalized['personal_contact'] = v
        else:
            normalized[k_str] = v
    return normalized

def format_date_dm4y(date_iso):
    if not date_iso:
        return ""
    try:
        d = datetime.datetime.strptime(str(date_iso)[:10], '%Y-%m-%d').date()
        return d.strftime('%d/%m/%Y')
    except Exception:
        return str(date_iso)

def calculate_age(dob_iso):
    if not dob_iso:
        return ""
    try:
        dob = datetime.date.fromisoformat(str(dob_iso)[:10])
        today = datetime.date.today()
        age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        return age
    except Exception:
        return ""

@app.route('/api/employees', methods=['GET', 'POST'])
@login_required
def handle_employees():
    tenant_id = g.current_user["tenant_id"]
    if request.method == 'POST':
        data = request.json or {}
        emp_id = data.get('emp_id')
        name = data.get('name')
        email = data.get('email')
        password = data.get('password') or 'password123'
        designation = data.get('designation', 'Employee')
        department = data.get('department', 'General')
        dob = parse_date_dm4y(data.get('dob'))
        doj = parse_date_dm4y(data.get('doj'))
        personal_email = data.get('personal_email', '')
        current_address = data.get('current_address', '')
        office_contact = data.get('office_contact', '')
        personal_contact = data.get('personal_contact', '')
        
        # No compulsory fields fallback
        if not name:
            name = ""
        if not email or str(email).strip() == "":
            email = ""
        else:
            email = str(email).strip()
            
        try:
            if email != "":
                email_domain = email.split('@')[-1] if '@' in email else ''
                if email_domain.lower() != "semcogroups.com":
                    return jsonify({"detail": "Only semcogroups.com domain accounts are permitted."}), 400
                    
                # Check if this email is used by an admin
                admin_user = db.employees.find_one({"email": email, "role": "Admin (HR)"})
                if admin_user:
                    # Admin email: check if a regular employee profile already exists under this email
                    emp_profile = db.employees.find_one({"email": email, "role": {"$ne": "Admin (HR)"}})
                    if emp_profile:
                        return jsonify({"detail": "Only one employee profile can be created for the admin email."}), 400
                else:
                    # Regular email: check if any profile already exists
                    existing = db.employees.find_one({"email": email})
                    if existing:
                        return jsonify({"detail": "Employee with this email already exists."}), 400
                
            res = db.employees.insert_one({
                "tenant_id": tenant_id,
                "emp_id": emp_id,
                "name": name,
                "email": email,
                "password": password,
                "designation": designation,
                "role": designation,
                "department": department,
                "dob": dob,
                "doj": doj,
                "birthday": dob,
                "joining_date": doj,
                "anniversary": doj,
                "personal_email": personal_email,
                "current_address": current_address,
                "office_contact": office_contact,
                "personal_contact": personal_contact,
                "status": "ACTIVE",
                "system_access_revoked": 0,
                "allow_personal_email_access": False
            })

            # Auto-create a blank document vault record for this new employee
            db.employee_documents.update_one(
                {"employee_id": res.inserted_id, "tenant_id": tenant_id},
                {"$setOnInsert": {
                    "employee_id": res.inserted_id,
                    "tenant_id": tenant_id,
                    "personal_documents": {},
                    "company_documents": [],
                    "created_at": datetime.datetime.utcnow().isoformat()
                }},
                upsert=True
            )

            check_and_send_celebrations()
            emp = db.employees.find_one({"_id": res.inserted_id})
            return jsonify({
                "id": str(emp["_id"]),
                "emp_id": emp.get("emp_id", ""),
                "name": emp["name"],
                "email": emp["email"],
                "designation": emp.get("designation", emp.get("role", "Employee")),
                "department": emp["department"],
                "status": emp["status"]
            })
        except Exception as e:
            return jsonify({"detail": str(e)}), 500
    else:
        try:
            active_only = request.args.get('active_only', 'false').lower() == 'true'
            query = {
                "tenant_id": tenant_id
            }
            if active_only:
                query["status"] = "ACTIVE"
                
            employees = list(db.employees.find(query))
            return jsonify([{
                "id": str(e["_id"]),
                "emp_id": e.get("emp_id") or "",
                "name": e.get("name") or "",
                "email": e.get("email") or "",
                "designation": e.get("designation") or e.get("role") or "",
                "role": e.get("role") or e.get("designation") or "",
                "department": e.get("department") or "",
                "dob": format_date_dm4y(e.get("dob") or e.get("birthday") or ""),
                "doj": format_date_dm4y(e.get("doj") or e.get("joining_date") or e.get("anniversary") or ""),
                "age": calculate_age(e.get("dob") or e.get("birthday") or ""),
                "personal_email": e.get("personal_email") or "",
                "current_address": e.get("current_address") or "",
                "office_contact": e.get("office_contact") or "",
                "personal_contact": e.get("personal_contact") or "",
                "status": e.get("status") or "ACTIVE",
                "allow_personal_email_access": e.get("allow_personal_email_access", False)
            } for idx, e in enumerate(employees)])
        except Exception as e:
            return jsonify({"detail": str(e)}), 500

@app.route('/api/employees/import', methods=['POST'])
@login_required
def import_employees():
    if 'file' not in request.files:
        return jsonify({"detail": "No file uploaded"}), 400
        
    file = request.files['file']
    filename = file.filename or ''
    tenant_id = g.current_user["tenant_id"]
    
    imported_count = 0
    try:
        rows = []
        filename_lower = filename.lower()
        if filename_lower.endswith('.csv'):
            csv_data = file.stream.read().decode('utf-8').splitlines()
            reader = csv.DictReader(csv_data)
            for r in reader:
                rows.append(r)
        elif filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(file.stream, data_only=True)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            for row_idx in range(2, sheet.max_row + 1):
                row_values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers) + 1)]
                if not any(row_values):
                    continue
                row_dict = dict(zip(headers, row_values))
                rows.append(row_dict)
        else:
            return jsonify({"detail": "Unsupported file format. Use CSV or XLSX."}), 400
            
        print(f"[IMPORT] Parsed {len(rows)} raw rows from file {filename}.")
        for r in rows:
            r = normalize_keys(r)
            emp_id = str(r.get('emp_id')).strip() if r.get('emp_id') is not None else ''
            name = r.get('name')
            email = r.get('email')
            password = str(r.get('password')).strip() if r.get('password') is not None else 'password123'
            designation = str(r.get('designation')).strip() if r.get('designation') is not None else ''
            department = str(r.get('department')).strip() if r.get('department') is not None else ''
            dob = parse_date_dm4y(r.get('dob'))
            doj = parse_date_dm4y(r.get('doj'))
            personal_email = str(r.get('personal_email')).strip() if r.get('personal_email') is not None else ''
            current_address = str(r.get('current_address')).strip() if r.get('current_address') is not None else ''
            office_contact = str(r.get('office_contact')).strip() if r.get('office_contact') is not None else ''
            personal_contact = str(r.get('personal_contact')).strip() if r.get('personal_contact') is not None else ''
            
            # Smart Fallbacks so no row is skipped
            if not name or str(name).strip() == "":
                name = "Employee"
            else:
                name = str(name).strip()
                
            # If email is empty, but personal_email contains a corporate email, swap them
            if not email or str(email).strip() == "":
                if personal_email and personal_email.lower().endswith("@semcogroups.com"):
                    email = personal_email
                    personal_email = ""
                    
            if not email or str(email).strip() == "":
                email = ""
            else:
                email = str(email).strip()
                if not email.lower().endswith("@semcogroups.com"):
                    if not personal_email or personal_email == "":
                        personal_email = email
                    email = ""

            is_new = True
            existing = None
            
            # Check if this email matches an existing user
            if email and email != "":
                existing = db.employees.find_one({"email": email})
                if existing:
                    is_new = False
            elif emp_id and emp_id != "":
                existing = db.employees.find_one({"emp_id": emp_id})
                if existing:
                    is_new = False
                    
            if is_new and email and email != "":
                # Ensure the email is unique in the database
                email_base = email
                counter = 1
                while db.employees.find_one({"email": email}):
                    username_part, domain_part = email_base.split('@')
                    email = f"{username_part}{counter}@{domain_part}"
                    counter += 1
                    
            if existing:
                db.employees.update_one({"_id": existing["_id"]}, {"$set": {
                    "emp_id": emp_id or existing.get("emp_id") or "",
                    "name": name,
                    "password": password,
                    "designation": designation,
                    "role": designation,
                    "department": department,
                    "dob": dob,
                    "doj": doj,
                    "birthday": dob,
                    "joining_date": doj,
                    "anniversary": doj,
                    "personal_email": personal_email,
                    "current_address": current_address,
                    "office_contact": office_contact,
                    "personal_contact": personal_contact,
                    "status": "ACTIVE"
                }})
            else:
                db.employees.insert_one({
                    "tenant_id": tenant_id,
                    "emp_id": emp_id,
                    "name": name,
                    "email": email,
                    "password": password,
                    "designation": designation,
                    "role": designation,
                    "department": department,
                    "dob": dob,
                    "doj": doj,
                    "birthday": dob,
                    "joining_date": doj,
                    "anniversary": doj,
                    "personal_email": personal_email,
                    "current_address": current_address,
                    "office_contact": office_contact,
                    "personal_contact": personal_contact,
                    "status": "ACTIVE",
                    "system_access_revoked": 0,
                    "allow_personal_email_access": False
                })
            imported_count += 1
            
        check_and_send_celebrations()
        print(f"[IMPORT] Successfully imported/synchronized {imported_count} employee records.")
        return jsonify({"imported": imported_count})
    except Exception as e:
        return jsonify({"detail": f"Failed to import: {str(e)}"}), 500

@app.route('/api/employees/<emp_id>/archive', methods=['PUT'])
@login_required
def archive_employee(emp_id):
    try:
        res = db.employees.update_one(
            {"_id": ObjectId(emp_id), "tenant_id": g.current_user["tenant_id"]},
            {"$set": {"status": "ARCHIVED"}}
        )
        if res.matched_count == 0:
            return jsonify({"detail": "Employee not found"}), 404
        return jsonify({"message": "Employee moved to archive successfully."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/employees/<emp_id>/restore', methods=['PUT'])
@login_required
def restore_employee(emp_id):
    try:
        res = db.employees.update_one(
            {"_id": ObjectId(emp_id), "tenant_id": g.current_user["tenant_id"]},
            {"$set": {"status": "ACTIVE"}}
        )
        if res.matched_count == 0:
            return jsonify({"detail": "Employee not found"}), 404
        return jsonify({"message": "Employee restored to active roster."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/employees/<emp_id>', methods=['DELETE', 'PUT'])
@login_required
def handle_single_employee(emp_id):
    if request.method == 'DELETE':
        try:
            res = db.employees.delete_one(
                {"_id": ObjectId(emp_id), "tenant_id": g.current_user["tenant_id"]}
            )
            if res.deleted_count == 0:
                return jsonify({"detail": "Employee not found"}), 404
            return jsonify({"message": "Employee permanently deleted."})
        except Exception as e:
            return jsonify({"detail": str(e)}), 500
    elif request.method == 'PUT':
        data = request.json or {}
        emp_id_val = data.get('emp_id')
        name = data.get('name')
        email = data.get('email')
        designation = data.get('designation', '')
        department = data.get('department', '')
        dob = parse_date_dm4y(data.get('dob'))
        doj = parse_date_dm4y(data.get('doj'))
        personal_email = data.get('personal_email', '')
        current_address = data.get('current_address', '')
        office_contact = data.get('office_contact', '')
        personal_contact = data.get('personal_contact', '')
        
        allow_personal_email_access = data.get('allow_personal_email_access')
        
        # No compulsory fields fallback
        if not name:
            name = ""
            
        try:
            existing_emp = db.employees.find_one({"_id": ObjectId(emp_id)})
            was_allowed = existing_emp.get("allow_personal_email_access", False) if existing_emp else False
            
            update_set = {
                "emp_id": emp_id_val,
                "name": name,
                "designation": designation,
                "role": designation,
                "department": department,
                "dob": dob,
                "doj": doj,
                "birthday": dob,
                "joining_date": doj,
                "anniversary": doj,
                "personal_email": personal_email,
                "current_address": current_address,
                "office_contact": office_contact,
                "personal_contact": personal_contact
            }
            
            if allow_personal_email_access is not None:
                new_allowed = bool(allow_personal_email_access)
                update_set["allow_personal_email_access"] = new_allowed
                
                # Transition to true: send notification email and promote PENDING employee to ACTIVE
                if new_allowed and not was_allowed:
                    if existing_emp and existing_emp.get("status") == "PENDING":
                        update_set["status"] = "ACTIVE"
                    to_email = personal_email or (existing_emp.get("personal_email") if existing_emp else "")
                    if to_email:
                        scheme = "https" if request.is_secure or "vercel" in request.headers.get("Host", "").lower() else "http"
                        host = request.headers.get("Host", "localhost:8000")
                        portal_link = f"{scheme}://{host}/?signup=true"
                        subject = "Portal Access Granted via Personal Email"
                        body = f"""
                        <html>
                            <body>
                                <div style="padding: 24px; font-family: Arial, sans-serif; line-height: 1.6; color: #1e293b;">
                                    <h2 style="color: #15803d; margin-top: 0;">Access Approved!</h2>
                                    <p>Hello <strong>{name or (existing_emp.get('name') if existing_emp else 'Employee')}</strong>,</p>
                                    <p>HR has approved your request to access the SEMCO Groups HR Portal using your personal email address.</p>
                                    <p>You can now register/sign up and log in using your personal email: <strong>{to_email}</strong>.</p>
                                    <div style="margin: 24px 0;">
                                        <a href="{portal_link}" style="display: inline-block; padding: 12px 24px; background: linear-gradient(135deg, #15803d, #1d4ed8); color: white; text-decoration: none; border-radius: 8px; font-weight: bold;">
                                            Go to Portal Sign Up
                                        </a>
                                    </div>
                                    <p>Best regards,<br>HR Operations Team<br><strong>SEMCO Groups</strong></p>
                                </div>
                            </body>
                        </html>
                        """
                        send_email(to_email, subject, body)
            
            if email:
                email = str(email).strip()
                if email != "":
                    # Check if personal email access is allowed for this domain bypass
                    curr_allowed = update_set.get("allow_personal_email_access", was_allowed)
                    is_personal_allowed = (
                        curr_allowed or 
                        (email == personal_email) or 
                        (existing_emp and email == existing_emp.get("personal_email"))
                    )
                    if not email.lower().endswith("@semcogroups.com") and not is_personal_allowed:
                        return jsonify({"detail": "Only semcogroups.com domain accounts are permitted."}), 400
                    update_set["email"] = email
                    # Auto-promote PENDING employee to ACTIVE when company email is provisioned
                    if existing_emp and existing_emp.get("status") == "PENDING":
                        update_set["status"] = "ACTIVE"
                else:
                    update_set["email"] = ""
            else:
                update_set["email"] = ""

            res = db.employees.update_one(
                {"_id": ObjectId(emp_id), "tenant_id": g.current_user["tenant_id"]},
                {"$set": update_set}
            )
            if res.matched_count == 0:
                return jsonify({"detail": "Employee not found"}), 404
            check_and_send_celebrations()
            return jsonify({"message": "Employee updated successfully."})
        except Exception as e:
            return jsonify({"detail": str(e)}), 500

# --- Celebrations ---
def try_parse_date(date_str):
    if not date_str:
        return None
    date_str = str(date_str).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(date_str[:10], fmt).date()
        except ValueError:
            continue
    return None

@app.route('/api/celebrations/match', methods=['GET'])
@login_required
def get_celebrations_matches():
    try:
        tenant_id = g.current_user["tenant_id"]
        today = datetime.date.today()
        
        employees = list(db.employees.find({"tenant_id": tenant_id, "status": "ACTIVE"}))
        
        birthdays = []
        anniversaries = []
        
        for e in employees:
            birthday_val = e.get("dob") or e.get("birthday")
            if birthday_val:
                b_date = try_parse_date(birthday_val)
                if b_date and b_date.month == today.month and b_date.day == today.day:
                    birthdays.append({"id": str(e["_id"]), "name": e.get("name") or "Employee", "email": e.get("email") or ""})
                    
            anniversary_val = e.get("doj") or e.get("anniversary") or e.get("joining_date")
            if anniversary_val:
                a_date = try_parse_date(anniversary_val)
                if a_date and a_date.month == today.month and a_date.day == today.day:
                    anniversaries.append({
                        "id": str(e["_id"]),
                        "name": e.get("name") or "Employee",
                        "email": e.get("email") or "",
                        "joining_date": format_date_dm4y(anniversary_val)
                    })
                
        return jsonify({"birthdays": birthdays, "anniversaries": anniversaries})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/celebrations/send', methods=['POST'])
@login_required
def send_celebration_email():
    data = request.json or {}
    employee_id = data.get('employee_id')
    celebration_type = data.get('type') # BIRTHDAY or ANNIVERSARY
    
    if not employee_id or not celebration_type:
        return jsonify({"detail": "employee_id and type are required"}), 400
        
    try:
        emp = db.employees.find_one({"_id": ObjectId(employee_id)})
        if not emp:
            return jsonify({"detail": "Employee not found"}), 404
            
        dept = emp.get("department") or "General"
        desg = emp.get("designation") or emp.get("role") or "Employee"
        
        if celebration_type == 'BIRTHDAY':
            full_name = emp.get("name", "Employee")
            subject = f"Happy Birthday, {full_name}! 🎂🎉"
            body = f"""
            <html>
                <body style="font-family: Arial, sans-serif; background-color: #fdf2f8; padding: 20px; color: #1e293b;">
                    <div style="background-color: white; padding: 40px 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); text-align: center;">
                        <h2 style="color: #ec4899; text-align: center; margin-bottom: 8px;">Happy Birthday, {full_name}! 🎂🎈</h2>
                        <p style="font-size: 16px; line-height: 1.6; color: #475569; text-align: center; margin: 0 auto 20px; max-width: 480px;">
                            On behalf of the entire team, we wish you a fantastic birthday filled with joy, laughter, and success. Thank you for your amazing contributions to our organization!
                        </p>
                        <div style="text-align: center; font-size: 50px; margin: 24px 0;">🎉🎂🎁✨</div>
                        <p style="font-size: 14px; text-align: center; color: #94a3b8; margin-top: 24px;">Warmest wishes,<br>The People Operations Team</p>
                    </div>
                </body>
            </html>
            """

            # Broadcast to others
            employees = list(db.employees.find({"status": "ACTIVE"}))
            other_emails = []
            for other in employees:
                if str(other["_id"]) != str(emp["_id"]):
                    if other.get("email"):
                        other_emails.append(other["email"])
            if other_emails:
                broadcast_subject = f"Let's Celebrate {full_name}'s Birthday! 🥳"
                broadcast_body = f"""
                <html>
                    <body style="font-family: Arial, sans-serif; background-color: #fdf2f8; padding: 20px; color: #1e293b;">
                        <div style="background-color: white; padding: 40px 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); text-align: center;">
                            <h2 style="color: #ec4899; text-align: center; margin-bottom: 8px;">Let's Celebrate {full_name}'s Birthday! 🥳</h2>
                            <p style="font-size: 16px; line-height: 1.6; color: #475569; text-align: center; margin: 0 auto 20px; max-width: 480px;">
                                Today is a special day! Please join us in wishing a very Happy Birthday to our colleague, <strong>{full_name}</strong>.
                            </p>
                            <p style="font-size: 15px; color: #475569; text-align: center; margin: 6px 0;"><strong>Department:</strong> {dept}</p>
                            <p style="font-size: 15px; color: #475569; text-align: center; margin: 6px 0;"><strong>Designation:</strong> {desg}</p>
                            <div style="text-align: center; font-size: 50px; margin: 24px 0;">🎉🎂🎁✨</div>
                            <p style="font-size: 14px; text-align: center; color: #94a3b8; margin-top: 24px;">Warmest wishes,<br>The People Operations Team</p>
                        </div>
                    </body>
                </html>
                """
                send_email(", ".join(other_emails), broadcast_subject, broadcast_body)
        else:
            joining_val = emp.get("doj") or emp.get("joining_date") or emp.get("anniversary")
            joining_date_parsed = try_parse_date(joining_val) if joining_val else None
            joining_year = joining_date_parsed.year if joining_date_parsed else datetime.date.today().year
            years = datetime.date.today().year - joining_year
            if years <= 0:
                years = 1
            from .utils.scheduler import get_ordinal_suffix
            ordinal = get_ordinal_suffix(years)
            full_name = emp.get("name", "Employee")

            subject = f"Happy {ordinal} Work Anniversary, {full_name}! 🌟💼"
            body = f"""
            <html>
                <body style="font-family: Arial, sans-serif; background-color: #ecfdf5; padding: 20px; color: #1e293b;">
                    <div style="background-color: white; padding: 40px 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); text-align: center;">
                        <h2 style="color: #10b981; text-align: center; margin-bottom: 8px;">Happy {ordinal} Work Anniversary! 🌟</h2>
                        <p style="font-size: 18px; font-weight: bold; color: #065f46; text-align: center; margin: 0 0 16px;">🎊 Congratulations! 🎊</p>
                        <p style="font-size: 16px; line-height: 1.6; color: #475569; text-align: center; margin: 0 auto 20px; max-width: 480px;">
                            <b>{full_name}</b>, on completing your <b>{ordinal}</b> year with us! Thank you for your dedication, hard work, and support. We are proud to have you on our team.
                        </p>
                        <div style="text-align: center; font-size: 50px; margin: 24px 0;">💼✨🚀🏆</div>
                        <p style="font-size: 14px; text-align: center; color: #94a3b8; margin-top: 24px;">Best regards,<br>The People Operations Team</p>
                    </div>
                </body>
            </html>
            """

            # Broadcast to others
            employees = list(db.employees.find({"status": "ACTIVE"}))
            other_emails = []
            for other in employees:
                if str(other["_id"]) != str(emp["_id"]):
                    if other.get("email"):
                        other_emails.append(other["email"])
            if other_emails:
                broadcast_subject = f"Celebrating {full_name}'s {ordinal} Work Anniversary! 🚀"
                broadcast_body = f"""
                <html>
                    <body style="font-family: Arial, sans-serif; background-color: #ecfdf5; padding: 20px; color: #1e293b;">
                        <div style="background-color: white; padding: 40px 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); text-align: center;">
                            <h2 style="color: #10b981; text-align: center; margin-bottom: 8px;">Celebrating {full_name}'s {ordinal} Work Anniversary! 🚀</h2>
                            <p style="font-size: 18px; font-weight: bold; color: #065f46; text-align: center; margin: 0 0 16px;">🎊 Congratulations! 🎊</p>
                            <p style="font-size: 16px; line-height: 1.6; color: #475569; text-align: center; margin: 0 auto 20px; max-width: 480px;">
                                Please join us in congratulating <strong>{full_name}</strong> on completing their <strong>{ordinal}</strong> year with SEMCO Groups!
                            </p>
                            <p style="font-size: 15px; color: #475569; text-align: center; margin: 6px 0;"><strong>Department:</strong> {dept}</p>
                            <p style="font-size: 15px; color: #475569; text-align: center; margin: 6px 0;"><strong>Designation:</strong> {desg}</p>
                            <div style="text-align: center; font-size: 50px; margin: 24px 0;">💼✨🚀🏆</div>
                            <p style="font-size: 14px; text-align: center; color: #94a3b8; margin-top: 24px;">Best regards,<br>The People Operations Team</p>
                        </div>
                    </body>
                </html>
                """
                send_email(", ".join(other_emails), broadcast_subject, broadcast_body)
            
        emails_to_send = [emp.get("email")]
        to_email_str = ", ".join([email for email in emails_to_send if email])
        if to_email_str:
            send_email(to_email_str, subject, body)
            
        return jsonify({"message": "Celebration email sent successfully!"})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/celebrations/all', methods=['GET'])
@login_required
def get_all_celebrations():
    try:
        tenant_id = g.current_user["tenant_id"]
        employees = list(db.employees.find({"tenant_id": tenant_id, "status": "ACTIVE"}))
        
        celebrations = []
        for e in employees:
            # parse dob
            dob_val = e.get("dob") or e.get("birthday")
            if dob_val:
                b_date = try_parse_date(dob_val)
                if b_date:
                    celebrations.append({
                        "id": str(e["_id"]),
                        "name": e.get("name") or "Employee",
                        "email": e.get("email") or "",
                        "personal_email": e.get("personal_email") or "",
                        "department": e.get("department") or "",
                        "designation": e.get("designation") or e.get("role") or "",
                        "type": "BIRTHDAY",
                        "month": b_date.month,
                        "day": b_date.day,
                        "date_str": dob_val
                    })
            # parse doj
            doj_val = e.get("doj") or e.get("anniversary") or e.get("joining_date")
            if doj_val:
                a_date = try_parse_date(doj_val)
                if a_date:
                    celebrations.append({
                        "id": str(e["_id"]),
                        "name": e.get("name") or "Employee",
                        "email": e.get("email") or "",
                        "personal_email": e.get("personal_email") or "",
                        "department": e.get("department") or "",
                        "designation": e.get("designation") or e.get("role") or "",
                        "type": "ANNIVERSARY",
                        "month": a_date.month,
                        "day": a_date.day,
                        "date_str": doj_val
                    })
        return jsonify(celebrations)
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Payroll Hub ---
@app.route('/api/payroll', methods=['GET'])
@login_required
def get_payrolls():
    try:
        tenant_id = g.current_user["tenant_id"]
        employees = list(db.employees.find({"tenant_id": tenant_id}))
        emp_map = {str(e["_id"]): e for e in employees}
        emp_ids = [e["_id"] for e in employees]
        
        records = list(db.payrolls.find({"employee_id": {"$in": emp_ids}}))
        
        result = []
        for r in records:
            emp_info = emp_map.get(str(r["employee_id"]), {})
            result.append({
                "id": str(r["_id"]),
                "employee_id": str(r["employee_id"]),
                "base_salary": r["base_salary"],
                "allowances": r["allowances"],
                "deductions": r["deductions"],
                "net_salary": r["net_salary"],
                "pay_period": r["pay_period"],
                "status": r["status"],
                "employee": {
                    "name": emp_info.get("name", ""),
                    "email": emp_info.get("email", "")
                }
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/payroll/upload', methods=['POST'])
@login_required
def upload_payroll_document():
    if 'file' not in request.files:
        return jsonify({"detail": "No file uploaded"}), 400
        
    file = request.files['file']
    filename = file.filename.lower()
    
    if not filename.endswith(('.csv', '.xlsx', '.xls', '.ods')):
        return jsonify({"detail": "Unsupported document format. Supported formats: .csv, .xlsx, .xls, .ods"}), 400
        
    tenant_id = g.current_user["tenant_id"]
    imported_count = 0
    try:
        import pandas as pd
        import io
        import math
        
        file_bytes = file.stream.read()
        
        if filename.endswith('.csv'):
            try:
                df = pd.read_csv(io.BytesIO(file_bytes), encoding='utf-8')
            except Exception:
                df = pd.read_csv(io.BytesIO(file_bytes), encoding='latin-1')
        else:
            df = pd.read_excel(io.BytesIO(file_bytes))
            
        # Clean columns and replace nulls
        df.columns = [str(c).strip() for c in df.columns]
        df = df.where(pd.notnull(df), None)
        
        rows = df.to_dict(orient='records')
        
        for row in rows:
            normalized_row = {str(k).strip().lower(): v for k, v in row.items()}
            
            def get_row_value(keys_list, default=None):
                for k in keys_list:
                    if k.lower() in normalized_row:
                        return normalized_row[k.lower()]
                return default

            email = get_row_value(['email', 'email id', 'employee email', 'company email', 'corporate email'])
            pay_period = get_row_value(['pay_period', 'pay period', 'period', 'month'])
            
            if not email or not pay_period:
                continue
                
            email = str(email).strip()
            pay_period = str(pay_period).strip()
            
            def get_float(val):
                if val is None or (isinstance(val, float) and math.isnan(val)):
                    return 0.0
                val_str = str(val).strip()
                if not val_str or val_str == '-':
                    return 0.0
                try:
                    return float(val_str.replace(',', ''))
                except ValueError:
                    return 0.0
                    
            def get_int(val):
                if val is None or (isinstance(val, float) and math.isnan(val)):
                    return 0
                val_str = str(val).strip()
                if not val_str or val_str == '-':
                    return 0
                try:
                    return int(float(val_str))
                except ValueError:
                    return 0
            
            base_salary = get_float(get_row_value(['base_salary', 'Basic Salary', 'basic_salary', 'base salary']))
            allowances = get_float(get_row_value(['allowances', 'total allowances', 'allowance']))
            deductions = get_float(get_row_value(['deductions', 'total deductions', 'deduction']))
            
            basic_salary = get_float(get_row_value(['Basic Salary', 'basic_salary', 'basic salary']) or base_salary)
            hra = get_float(get_row_value(['HRA', 'hra']))
            special_allowance = get_float(get_row_value(['Special Allowance', 'special_allowance', 'special allowance']))
            other_allowance = get_float(get_row_value(['Other Allowance', 'other_allowance', 'other allowance']))
            conveyance_allowance = get_float(get_row_value(['Conveyance Allowance', 'conveyance_allowance', 'conveyance allowance']))
            reimbursment = get_float(get_row_value(['Reimbursment', 'reimbursment', 'Reimbursement', 'reimbursement']))
            
            advance_decucted = get_float(get_row_value(['Advance Decucted', 'advance_decucted', 'Advance Deducted', 'advance_deducted']))
            mlwf = get_float(get_row_value(['MLWF', 'mlwf']))
            pf = get_float(get_row_value(['PF', 'pf']))
            esi = get_float(get_row_value(['ESI', 'esi']))
            pt = get_float(get_row_value(['PT', 'pt']))
            
            present_days = get_int(get_row_value(['Present Days', 'present_days', 'present days', 'attendance']))
            leaves_taken = get_int(get_row_value(['Leaves Taken', 'leaves_taken', 'leaves taken']))
            leaves_balance = get_int(get_row_value(['Leaves Balance', 'leaves_balance', 'leaves balance']))
            
            # If totals are 0 but individual items exist, sum them
            if base_salary == 0 and basic_salary > 0:
                base_salary = basic_salary
            if allowances == 0:
                allowances = hra + special_allowance + other_allowance + conveyance_allowance + reimbursment
            if deductions == 0:
                deductions = advance_decucted + mlwf + pf + esi + pt
                
            net_salary = base_salary + allowances - deductions
            
            emp = db.employees.find_one({"email": email, "tenant_id": tenant_id})
            if not emp:
                name_val = get_row_value(['employee name', 'name', 'employee_name', 'employee name'])
                if not name_val:
                    name_val = email.split('@')[0].replace('.', ' ').replace('_', ' ').title()
                
                emp_code = get_row_value(['employee code', 'code', 'employee_code', 'emp_id', 'emp id'])
                if not emp_code:
                    import random
                    emp_code = f"EMP{random.randint(1000, 9999)}"
                    
                desig = get_row_value(['designation', 'role']) or "Operations Associate"
                uan = get_row_value(['uan no', 'uan', 'uan_no', 'uan no.']) or "-"
                esic = get_row_value(['esic no', 'esic', 'esic_no', 'esic no.']) or "-"
                
                new_emp = {
                    "name": name_val,
                    "email": email,
                    "tenant_id": tenant_id,
                    "status": "ACTIVE",
                    "role": desig,
                    "department": "Operations",
                    "emp_id": emp_code,
                    "designation": desig,
                    "uan_no": uan,
                    "esic_no": esic,
                    "personal_email": email.split('@')[0] + "@gmail.com"
                }
                
                insert_res = db.employees.insert_one(new_emp)
                emp = db.employees.find_one({"_id": insert_res.inserted_id})
                
            emp_id = emp["_id"]
            
            payload = {
                "base_salary": base_salary,
                "allowances": allowances,
                "deductions": deductions,
                "net_salary": net_salary,
                "basic_salary": basic_salary,
                "hra": hra,
                "special_allowance": special_allowance,
                "other_allowance": other_allowance,
                "conveyance_allowance": conveyance_allowance,
                "reimbursment": reimbursment,
                "advance_decucted": advance_decucted,
                "mlwf": mlwf,
                "pf": pf,
                "esi": esi,
                "pt": pt,
                "present_days": present_days,
                "leaves_taken": leaves_taken,
                "leaves_balance": leaves_balance,
                "pay_period": pay_period,
                "status": "PENDING"
            }
            
            existing = db.payrolls.find_one({"employee_id": emp_id, "pay_period": pay_period})
            if existing:
                db.payrolls.update_one({"_id": existing["_id"]}, {"$set": payload})
            else:
                payload["employee_id"] = emp_id
                db.payrolls.insert_one(payload)
            imported_count += 1
            
        return jsonify({"imported": imported_count})
    except Exception as e:
        return jsonify({"detail": f"Document processing error: {str(e)}"}), 500

@app.route('/api/payroll/<payroll_id>/email', methods=['POST'])
@login_required
def email_encrypted_payslip(payroll_id):
    try:
        pr = db.payrolls.find_one({"_id": ObjectId(payroll_id)})
        if not pr:
            return jsonify({"detail": "Payroll record not found"}), 404
            
        emp = db.employees.find_one({"_id": pr["employee_id"]})
        if not emp:
            return jsonify({"detail": "Associated employee not found"}), 404
            
        temp_pdf = f"temp_payslip_{str(pr['_id'])}.pdf"
        output_pdf = f"encrypted_payslip_{str(pr['_id'])}.pdf"
        
        employee_data = {
            "name": emp.get("name", "Valued Employee"),
            "email": emp.get("email", ""),
            "role": emp.get("role", ""),
            "department": emp.get("department", ""),
            "status": emp.get("status", ""),
            "emp_id": emp.get("emp_id", "-"),
            "designation": emp.get("designation") or emp.get("role") or "-",
            "uan_no": emp.get("uan_no") or emp.get("uan") or "-",
            "esic_no": emp.get("esic_no") or emp.get("esic") or "-"
        }
        payroll_data = {
            "pay_period": pr.get("pay_period", ""),
            "present_days": pr.get("present_days", 30),
            "leaves_taken": pr.get("leaves_taken", 0),
            "leaves_balance": pr.get("leaves_balance", 0),
            "basic_salary": pr.get("basic_salary", pr.get("base_salary", 0.0)),
            "hra": pr.get("hra", 0.0),
            "special_allowance": pr.get("special_allowance", 0.0),
            "other_allowance": pr.get("other_allowance", 0.0),
            "conveyance_allowance": pr.get("conveyance_allowance", 0.0),
            "reimbursment": pr.get("reimbursment", 0.0),
            "advance_decucted": pr.get("advance_decucted", 0.0),
            "mlwf": pr.get("mlwf", 0.0),
            "pf": pr.get("pf", 0.0),
            "esi": pr.get("esi", 0.0),
            "pt": pr.get("pt", 0.0),
            "base_salary": pr.get("base_salary", 0.0),
            "allowances": pr.get("allowances", 0.0),
            "deductions": pr.get("deductions", 0.0)
        }
        
        generate_salary_breakup_pdf(employee_data, payroll_data, temp_pdf)
        
        # Encrypt with AES-256 using employee email
        password = emp["email"]
        encrypt_pdf_aes(temp_pdf, output_pdf, password)
        
        subject = f"Encrypted Payslip - Period {pr['pay_period']}"
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; background-color: #f8fafc; padding: 20px; color: #1e293b;">
                <div style="background-color: white; padding: 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">
                    <h2 style="color: #10b981; border-bottom: 2px solid #ecfdf5; padding-bottom: 10px; margin-top: 0;">Payslip Received 👔</h2>
                    <p style="font-size: 15px; color: #475569;">
                        Dear {emp['name']},<br><br>
                        Please find attached your salary breakup / payslip for the pay period <b>{pr['pay_period']}</b>.
                    </p>
                    <div style="background-color: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; margin: 20px 0; border-radius: 6px;">
                        <p style="margin: 0; font-size: 13px; color: #78350f;">
                            <strong>🔒 Password Protected PDF:</strong><br>
                            This document is encrypted with AES-256 for privacy. Use <b>your work email address</b> (e.g. <code>{emp['email']}</code>) to unlock it.
                        </p>
                    </div>
                    <p style="font-size: 13px; color: #94a3b8;">This is an automated operational dispatch. Please contact HR for any salary queries.</p>
                </div>
            </body>
        </html>
        """
        
        send_email(emp["email"], subject, body, attachment_path=output_pdf, attachment_name=f"payslip_{pr['pay_period']}.pdf")
        
        if os.path.exists(output_pdf):
            try:
                os.remove(output_pdf)
            except OSError:
                pass
                
        db.payrolls.update_one({"_id": ObjectId(payroll_id)}, {"$set": {"status": "SENT"}})
        return jsonify({"message": "Payslip emailed successfully!"})
    except Exception as e:
        return jsonify({"detail": f"Payslip generation failed: {str(e)}"}), 500

# --- CTC Export ---
@app.route('/api/payroll/ctc/export', methods=['POST'])
@login_required
def export_ctc():
    import tempfile, datetime
    try:
        data = request.json or {}
        fmt = data.get('format', 'pdf').lower()   # 'pdf', 'excel', 'word'
        employee_info = data.get('employee_info', {})
        ctc_data = data.get('ctc_data', {})
        date_str = datetime.datetime.now().strftime("%d %B %Y")
        location = data.get('location', 'Pune')

        suffix_map = {'pdf': '.pdf', 'excel': '.xlsx', 'word': '.docx'}
        suffix = suffix_map.get(fmt, '.pdf')

        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp_path = tmp.name

        if fmt == 'pdf':
            generate_pdf(ctc_data, employee_info, location, tmp_path)
            mime = 'application/pdf'
            filename = 'CTC_Breakup.pdf'
        elif fmt == 'excel':
            generate_excel(ctc_data, employee_info, location, tmp_path)
            mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = 'CTC_Breakup.xlsx'
        else:
            generate_word(ctc_data, employee_info, location, tmp_path)
            mime = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            filename = 'CTC_Breakup.docx'

        from flask import send_file
        return send_file(
            tmp_path,
            mimetype=mime,
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        import traceback
        return jsonify({"detail": f"CTC export failed: {str(e)}", "trace": traceback.format_exc()}), 500

# --- Daily Pulse ---
@app.route('/api/daily-pulse/quotes', methods=['GET', 'POST'])
@login_required
def handle_quotes():
    try:
        if request.method == 'POST':
            data = request.json or {}
            text = data.get('text')
            author = data.get('author', 'Unknown')
            if not text:
                return jsonify({"detail": "text is required"}), 400
                
            res = db.quotes.insert_one({"text": text, "author": author})
            q = db.quotes.find_one({"_id": res.inserted_id})
            return jsonify(serialize_doc(q))
        else:
            quotes = list(db.quotes.find({}))
            return jsonify(serialize_list(quotes))
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/daily-pulse/today', methods=['GET'])
@login_required
def get_today_daily_pulse():
    try:
        tenant_id = g.current_user["tenant_id"]
        today = datetime.datetime.now().date().isoformat()
        pulse = db.daily_pulse_schedule.find_one({"date": today, "tenant_id": tenant_id})
        if pulse:
            return jsonify({
                "quote": pulse["quote"],
                "author": pulse.get("author", "Unknown"),
                "date": pulse["date"]
            })
            
        quotes = list(db.quotes.find({}))
        if not quotes:
            quotes = [
                {"text": "The only way to do great work is to love what you do.", "author": "Steve Jobs"},
                {"text": "Success is not final, failure is not fatal: it is the courage to continue that counts.", "author": "Winston Churchill"},
                {"text": "Believe you can and you're halfway there.", "author": "Theodore Roosevelt"},
                {"text": "Act as if what you do makes a difference. It does.", "author": "William James"}
            ]
            
        day_of_year = datetime.datetime.now().timetuple().tm_yday
        selected = quotes[day_of_year % len(quotes)]
        return jsonify({
            "quote": selected.get("text") or selected.get("quote"),
            "author": selected.get("author", "Unknown"),
            "date": today
        })
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/daily-pulse/schedule', methods=['GET'])
@login_required
def get_daily_pulse_schedule():
    try:
        tenant_id = g.current_user["tenant_id"]
        from .utils.scheduler import ensure_daily_pulse_schedule
        
        # Pre-populate schedule for current and next month
        today = datetime.datetime.now()
        ensure_daily_pulse_schedule(today.year, today.month)
        
        next_month_date = today + datetime.timedelta(days=32)
        ensure_daily_pulse_schedule(next_month_date.year, next_month_date.month)
        
        schedules = list(db.daily_pulse_schedule.find({"tenant_id": tenant_id}))
        
        # Dynamic active employees manifest (delta-sync)
        employees = list(db.employees.find({"tenant_id": tenant_id, "status": "ACTIVE"}))
        manifest = [{
            "name": emp.get("name") or "Employee",
            "email": emp.get("email") or ""
        } for emp in employees if emp.get("email")]
        
        result = []
        for s in schedules:
            result.append({
                "id": str(s["_id"]),
                "date": s["date"],
                "time": s.get("time", "09:00"),
                "quote": s["quote"],
                "author": s.get("author", "Unknown"),
                "status": s.get("status", "Scheduled"),
                "delivered_at": s.get("delivered_at"),
                "recipients": manifest
            })
            
        result.sort(key=lambda x: x["date"])
        return jsonify(result)
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/daily-pulse/trigger', methods=['POST'])
@login_required
def trigger_daily_pulse():
    try:
        tenant_id = g.current_user["tenant_id"]
        from .utils.scheduler import check_and_send_daily_pulse
        
        today = datetime.datetime.now().date().isoformat()
        pulse = db.daily_pulse_schedule.find_one({"date": today, "tenant_id": tenant_id})
        if pulse:
            check_and_send_daily_pulse()
            return jsonify({"message": "Today's scheduled Daily Pulse was dispatched successfully."})
            
        quotes = list(db.quotes.find({}))
        if not quotes:
            return jsonify({"detail": "No quotes available in the library."}), 400
            
        quote = random.choice(quotes)
        employees = list(db.employees.find({"tenant_id": tenant_id, "status": "ACTIVE"}))
        emails = [emp.get("email") for emp in employees if emp.get("email")]
        if not emails:
            return jsonify({"detail": "No active employee emails found."}), 400
            
        from .utils.mailer import send_email
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; background-color: #f8fafc; padding: 20px; color: #1e293b;">
                <div style="background-color: white; padding: 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">
                    <h2 style="color: #3b82f6; border-bottom: 2px solid #eff6ff; padding-bottom: 10px; margin-top: 0;">Daily Pulse 🌟</h2>
                    <blockquote style="font-size: 18px; font-style: italic; color: #334155; border-left: 4px solid #3b82f6; padding-left: 15px; margin: 20px 0;">
                        "{quote['text']}"
                    </blockquote>
                    <p style="text-align: right; font-weight: bold; color: #64748b;">— {quote.get('author', 'Unknown')}</p>
                    <hr style="border: 0; border-top: 1px solid #e2e8f0; margin: 20px 0;">
                    <p style="font-size: 12px; color: #94a3b8; text-align: center;">You received this as part of your company's daily motivation program.</p>
                </div>
            </body>
        </html>
        """
        to_emails_str = ", ".join(emails)
        send_email(to_emails_str, "Daily Pulse 🌟 (Manual Blast)", body)
        return jsonify({"message": f"Manual Daily Pulse blast sent successfully to {len(emails)} employees."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Email Sandbox ---
@app.route('/api/sandbox/test-email', methods=['POST'])
@login_required
def sandbox_test_email():
    try:
        data = request.json or {}
        email = data.get('email')
        name = data.get('name', 'Valued Employee')
        email_type = data.get('type', 'birthday')
        
        if not email:
            return jsonify({"detail": "Recipient email is required."}), 400
            
        from .utils.mailer import send_email
        
        subject = ""
        body = ""
        
        if email_type == 'birthday':
            subject = f"Happy Birthday, {name}! 🎂🎉"
            body = f"""
            <html>
                <body style="font-family: Arial, sans-serif; background-color: #fdf2f8; padding: 20px; color: #1e293b;">
                    <div style="background-color: white; padding: 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">
                        <h2 style="color: #ec4899; text-align: center;">Happy Birthday, {name}! 🎂🎈</h2>
                        <p style="font-size: 16px; line-height: 1.5; color: #475569;">
                            This is a sandboxed test email simulating a Birthday wish. We wish you a fantastic year filled with joy, laughter, and success. Thank you for your amazing contributions!
                        </p>
                        <div style="text-align: center; font-size: 40px; margin: 20px 0;">🎉🎂🎁✨</div>
                        <p style="font-size: 14px; text-align: center; color: #94a3b8;">Warmest wishes,<br>The People Operations Team (Sandbox Mode)</p>
                    </div>
                </body>
            </html>
            """
        elif email_type == 'anniversary':
            years = data.get('years', 3)
            subject = f"Happy Work Anniversary, {name}! 🌟💼"
            body = f"""
            <html>
                <body style="font-family: Arial, sans-serif; background-color: #ecfdf5; padding: 20px; color: #1e293b;">
                    <div style="background-color: white; padding: 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">
                        <h2 style="color: #10b981; text-align: center;">Happy Work Anniversary! 🌟</h2>
                        <p style="font-size: 16px; line-height: 1.5; color: #475569;">
                            This is a sandboxed test email simulating a Work Anniversary wish. Congratulations, <b>{name}</b>, on completing {years} years with us! Thank you for your hard work and support.
                        </p>
                        <div style="text-align: center; font-size: 40px; margin: 20px 0;">💼✨🚀🏆</div>
                        <p style="font-size: 14px; text-align: center; color: #94a3b8;">Best regards,<br>The People Operations Team (Sandbox Mode)</p>
                    </div>
                </body>
            </html>
            """
        elif email_type == 'pulse':
            quote = data.get('quote', 'The only way to do great work is to love what you do.')
            author = data.get('author', 'Steve Jobs')
            subject = "Daily Pulse 🌟"
            body = f"""
            <html>
                <body style="font-family: Arial, sans-serif; background-color: #f8fafc; padding: 20px; color: #1e293b;">
                    <div style="background-color: white; padding: 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">
                        <h2 style="color: #3b82f6; border-bottom: 2px solid #eff6ff; padding-bottom: 10px; margin-top: 0;">Daily Pulse 🌟</h2>
                        <p style="font-size: 14px; color: #94a3b8; margin-bottom: 15px;"><i>[Sandbox Test Dispatch]</i></p>
                        <blockquote style="font-size: 18px; font-style: italic; color: #334155; border-left: 4px solid #3b82f6; padding-left: 15px; margin: 20px 0;">
                            "{quote}"
                        </blockquote>
                        <p style="text-align: right; font-weight: bold; color: #64748b;">— {author}</p>
                        <hr style="border: 0; border-top: 1px solid #e2e8f0; margin: 20px 0;">
                        <p style="font-size: 12px; color: #94a3b8; text-align: center;">You received this as part of your company's daily motivation program.</p>
                    </div>
                </body>
            </html>
            """
        else:
            return jsonify({"detail": f"Unknown email type: {email_type}"}), 400
            
        send_email(email, subject, body)
        return jsonify({"message": f"Sandbox test email ({email_type}) sent successfully to {email}."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Surprise Ops ---
@app.route('/api/surprise-ops/coupons', methods=['GET', 'POST'])
@login_required
def handle_coupons():
    try:
        tenant_id = g.current_user["tenant_id"]
        if request.method == 'POST':
            data = request.json or {}
            code = data.get('code')
            amount = float(data.get('amount', 0))
            brand = data.get('brand')
            
            if not code or not brand:
                return jsonify({"detail": "code and brand are required"}), 400
                
            res = db.coupons.insert_one({
                "tenant_id": tenant_id,
                "code": code,
                "amount": amount,
                "brand": brand,
                "assigned_to_email": None,
                "is_redeemed": False,
                "is_archived": False
            })
            c = db.coupons.find_one({"_id": res.inserted_id})
            return jsonify(serialize_doc(c))
        else:
            coupons = list(db.coupons.find({"tenant_id": tenant_id}))
            return jsonify(serialize_list(coupons))
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/surprise-ops/coupons/<coupon_id>/send', methods=['POST'])
@login_required
def send_coupon(coupon_id):
    data = request.json or {}
    email = data.get('email')
    
    if not email:
        return jsonify({"detail": "email is required"}), 400
        
    try:
        c = db.coupons.find_one({"_id": ObjectId(coupon_id)})
        if not c:
            return jsonify({"detail": "Coupon not found"}), 404
            
        db.coupons.update_one({"_id": ObjectId(coupon_id)}, {"$set": {
            "assigned_to_email": email,
            "is_redeemed": True
        }})
        
        subject = "Surprise Reward Inside! 🎁✨"
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; background-color: #faf5ff; padding: 20px; color: #1e293b;">
                <div style="background-color: white; padding: 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">
                    <h2 style="color: #8b5cf6; text-align: center;">A Surprise Token of Appreciation! 🎁</h2>
                    <p style="font-size: 15px; color: #475569; text-align: center;">
                        You have received a special instant incentive voucher. Thank you for your great work!
                    </p>
                    <div style="background-color: #faf5ff; border: 2px dashed #8b5cf6; padding: 20px; margin: 25px 0; border-radius: 8px; text-align: center;">
                        <span style="font-size: 13px; text-transform: uppercase; color: #6b21a8; font-weight: bold;">{c["brand"]}</span><br>
                        <span style="font-size: 32px; font-weight: 800; color: #8b5cf6; margin: 10px 0; display: block;">₹{c["amount"]:.2f}</span>
                        <span style="font-size: 16px; font-family: monospace; letter-spacing: 1px; background: #f3e8ff; padding: 6px 15px; border-radius: 4px; display: inline-block;">{c["code"]}</span>
                    </div>
                    <p style="font-size: 13px; color: #94a3b8; text-align: center;">Redeem this directly on the merchant's portal. Enjoy!</p>
                </div>
            </body>
        </html>
        """
        send_email(email, subject, body)
        return jsonify({"message": "Coupon successfully sent!"})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/surprise-ops/coupons/<coupon_id>/archive', methods=['PUT'])
@login_required
def archive_coupon(coupon_id):
    try:
        tenant_id = g.current_user["tenant_id"]
        res = db.coupons.update_one(
            {"_id": ObjectId(coupon_id), "tenant_id": tenant_id},
            {"$set": {"is_archived": True}}
        )
        if res.matched_count == 0:
            return jsonify({"detail": "Coupon not found"}), 404
        return jsonify({"message": "Coupon successfully archived!"})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/surprise-ops/coupons/<coupon_id>', methods=['DELETE'])
@login_required
def delete_coupon(coupon_id):
    try:
        tenant_id = g.current_user["tenant_id"]
        res = db.coupons.delete_one(
            {"_id": ObjectId(coupon_id), "tenant_id": tenant_id}
        )
        if res.deleted_count == 0:
            return jsonify({"detail": "Coupon not found"}), 404
        return jsonify({"message": "Coupon successfully deleted forever!"})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Surprise Ops Appreciation ---
@app.route('/api/surprise-ops/appreciation', methods=['GET', 'POST'])
@login_required
def handle_appreciation():
    try:
        tenant_id = g.current_user["tenant_id"]
        if request.method == 'POST':
            employee_name = request.form.get('employee_name')
            employee_email = request.form.get('employee_email')
            date = request.form.get('date')
            reason = request.form.get('reason')
            a_type = request.form.get('type')
            
            import json
            recipients_json = request.form.get('announcement_recipients', '[]')
            try:
                announcement_recipients = json.loads(recipients_json)
            except Exception:
                announcement_recipients = []
            
            if not employee_name or not employee_email or not date or not reason or not a_type:
                return jsonify({"detail": "All fields are required"}), 400
                
            certificate_filename = None
            certificate_url = None
            
            if a_type == 'CERTIFICATE':
                if 'certificate' not in request.files:
                    return jsonify({"detail": "Certificate file is required for Certificate of Appreciation"}), 400
                file = request.files['certificate']
                if file.filename == '':
                    return jsonify({"detail": "No certificate file selected"}), 400
                
                from werkzeug.utils import secure_filename
                import os
                filename = secure_filename(file.filename)
                
                upload_dir = get_upload_dir('certificates')
                
                import uuid
                unique_filename = f"{uuid.uuid4().hex}_{filename}"
                filepath = os.path.join(upload_dir, unique_filename)
                file.save(filepath)
                
                certificate_filename = filename
                certificate_url = f"/static/uploads/certificates/{unique_filename}"
                
            res = db.appreciations.insert_one({
                "tenant_id": tenant_id,
                "employee_name": employee_name,
                "employee_email": employee_email,
                "date": date,
                "reason": reason,
                "type": a_type,
                "certificate_filename": certificate_filename,
                "certificate_url": certificate_url,
                "announcement_recipients": announcement_recipients,
                "status": "PENDING"
            })
            item = db.appreciations.find_one({"_id": res.inserted_id})
            return jsonify(serialize_doc(item))
        else:
            items = list(db.appreciations.find({"tenant_id": tenant_id}))
            return jsonify(serialize_list(items))
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/surprise-ops/appreciation/<appreciation_id>/send', methods=['POST'])
@login_required
def send_appreciation(appreciation_id):
    smtp_server = None
    try:
        from app.utils.mailer import SMTP_HOST, SMTP_USER, SMTP_PASSWORD, SMTP_PORT
        if SMTP_HOST and SMTP_USER and SMTP_PASSWORD:
            try:
                import smtplib
                smtp_server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=5)
                smtp_server.starttls()
                smtp_server.login(SMTP_USER, SMTP_PASSWORD)
            except Exception as smtp_err:
                print(f"[SMTP CON] Failed to pre-connect: {str(smtp_err)}")
                smtp_server = None
    except Exception:
        pass

    try:
        tenant_id = g.current_user["tenant_id"]
        appr = db.appreciations.find_one({"_id": ObjectId(appreciation_id), "tenant_id": tenant_id})
        if not appr:
            if smtp_server:
                try:
                    smtp_server.quit()
                except Exception:
                    pass
            return jsonify({"detail": "Appreciation record not found"}), 404
            
        email = appr["employee_email"]
        name = appr["employee_name"]
        a_type = appr["type"]
        reason = appr["reason"]
        date_str = appr["date"]
        
        # Resolve awardee's personal_email and profile photo if registered
        awardee = db.employees.find_one({"email": email, "tenant_id": tenant_id})
        personal_email = awardee.get("personal_email") if awardee else None
        
        photo_html = ""
        inline_images = None
        if a_type == 'MONTH' and awardee:
            photo_url = None
            doc_vault = db.employee_documents.find_one({"employee_id": awardee["_id"]})
            if doc_vault and doc_vault.get("personal_documents"):
                photo_url = doc_vault["personal_documents"].get("profile_photo_url")
            
            # Construct absolute URL if it is a relative path
            if photo_url and photo_url.startswith('/'):
                base_url = request.host_url.rstrip('/')
                if "vercel" in base_url or os.environ.get("VERCEL"):
                    base_url = base_url.replace("http://", "https://")
                photo_url = f"{base_url}{photo_url}"
            
            image_data = None
            subtype = 'png'
            if photo_url and photo_url.startswith('data:image/'):
                try:
                    header, encoded = photo_url.split(',', 1)
                    if 'base64' in header:
                        image_data = base64.b64decode(encoded)
                        if 'jpeg' in header or 'jpg' in header:
                            subtype = 'jpeg'
                except Exception:
                    pass
            
            if image_data:
                photo_url = "cid:profile_photo"
                inline_images = [{"content_id": "profile_photo", "data": image_data, "filename": f"profile.{subtype}"}]
            elif not photo_url or not photo_url.startswith('http'):
                import urllib.parse
                encoded_name = urllib.parse.quote_plus(name)
                photo_url = f"https://ui-avatars.com/api/?name={encoded_name}&background=8b5cf6&color=fff&size=200&rounded=true"
                
            photo_html = f"""
            <div style="text-align: center; margin: 25px 0;">
                <div style="display: inline-block; width: 140px; height: 140px; border-radius: 50%; overflow: hidden; border: 4px solid #8b5cf6; box-shadow: 0 4px 10px rgba(0,0,0,0.15);">
                    <img src="{photo_url}" alt="{name}" style="width: 100%; height: 100%; object-fit: cover; display: block;" />
                </div>
            </div>
            """
            
        subject = ""
        attachment_path = None
        attachment_name = None
        
        type_labels = {
            "CERTIFICATE": "Certificate of Appreciation",
            "CARD": "Job Well Done Card",
            "MONTH": "Employee of the Month"
        }
        type_label = type_labels.get(a_type, "Award of Excellence")
        
        if a_type == 'CERTIFICATE':
            subject = f"Congratulations on your Certificate of Appreciation! 📜🎓"
            if appr.get("certificate_url"):
                import os
                unique_filename = appr["certificate_url"].split('/')[-1]
                upload_dir = get_upload_dir('certificates')
                attachment_path = os.path.join(upload_dir, unique_filename)
                attachment_name = appr.get("certificate_filename", "Certificate.pdf")
        elif a_type == 'CARD':
            subject = f"Great Job! You received a Job Well Done Card 🌟🙌"
        else:
            subject = f"Fantastic! You have been named Employee of the Month 🏆🎖️"
            
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; background-color: #f5f3ff; padding: 20px; color: #1e293b;">
                <div style="background-color: white; padding: 40px; border-radius: 16px; max-width: 600px; margin: 0 auto; box-shadow: 0 10px 15px -3px rgba(0,0,0,0.1); border-top: 8px solid #8b5cf6;">
                    <div style="text-align: center; margin-bottom: 25px;">
                        <span style="font-size: 50px;">🏆</span>
                        <h2 style="color: #6d28d9; margin-top: 10px; font-size: 24px;">Honoring Excellence</h2>
                        <span style="font-size: 13px; text-transform: uppercase; letter-spacing: 1.5px; color: #8b5cf6; font-weight: bold; background: #f5f3ff; padding: 4px 12px; border-radius: 20px;">
                            {type_label}
                        </span>
                    </div>
                    
                    {photo_html}
                    
                    <p style="font-size: 16px; color: #334155; line-height: 1.6;">
                        Dear <b>{name}</b>,<br><br>
                        We are thrilled to present you with this recognition in appreciation of your hard work, dedication, and outstanding contributions to our team.
                    </p>
                    
                    <div style="background-color: #faf5ff; border: 1px solid #ddd6fe; padding: 25px; margin: 25px 0; border-radius: 12px;">
                        <h4 style="margin: 0 0 10px 0; color: #5b21b6; font-size: 15px; text-transform: uppercase; letter-spacing: 0.5px;">Award Citation</h4>
                        <p style="margin: 0; font-size: 15px; font-style: italic; color: #4c1d95; line-height: 1.5;">
                            "{reason}"
                        </p>
                    </div>
                    
                    <table style="width: 100%; border-top: 1px solid #e2e8f0; margin-top: 25px; padding-top: 20px; font-size: 13px; color: #64748b;">
                        <tr>
                            <td><b>Awarded Date:</b> {date_str}</td>
                            <td style="text-align: right;"><b>HR Operations Team</b></td>
                        </tr>
                    </table>
                </div>
            </body>
        </html>
        """
        
        # 1. Send the appreciation card to the awardee's mailboxes
        send_email(email, subject, body, attachment_path=attachment_path, attachment_name=attachment_name, inline_images=inline_images, server=smtp_server)
            
        # 2. Dispatch announcement emails to selected employees
        announcement_recipients = appr.get("announcement_recipients") or []
        if announcement_recipients:
            rec_object_ids = []
            for rid in announcement_recipients:
                try:
                    rec_object_ids.append(ObjectId(rid))
                except Exception:
                    pass
            
            recipients = list(db.employees.find({"_id": {"$in": rec_object_ids}}))
            
            dept = awardee.get("department", "General") if awardee else "General"
            desg = (awardee.get("designation") or awardee.get("role") or "Employee") if awardee else "Employee"
            
            announce_subject = f"Company Announcement: Celebrating Employee Milestones! 🌟🎉"
            announce_body = f"""
            <html>
                <body style="font-family: Arial, sans-serif; background-color: #faf5ff; padding: 20px; color: #1e293b;">
                    <div style="background-color: white; padding: 40px; border-radius: 16px; max-width: 600px; margin: 0 auto; box-shadow: 0 10px 15px -3px rgba(0,0,0,0.1); border-top: 8px solid #8b5cf6; text-align: center;">
                        <span style="font-size: 50px;">🎉🌟</span>
                        <h2 style="color: #6d28d9; margin-top: 15px; font-size: 22px;">Celebrating Employee Milestone!</h2>
                        
                        {photo_html}
                        
                        <p style="font-size: 15px; color: #475569; line-height: 1.6; text-align: left; margin-top: 20px;">
                            Dear Team,<br><br>
                            We are proud to announce that <b>{name}</b> has been awarded the <b>{type_label}</b> on <b>{date_str}</b>.
                        </p>
                        <p style="font-size: 15px; color: #475569; line-height: 1.6; text-align: left; margin: 6px 0;">
                            <strong>Department:</strong> {dept}<br>
                            <strong>Designation:</strong> {desg}
                        </p>
                        <div style="background-color: #faf5ff; border: 1px solid #ddd6fe; padding: 20px; margin: 20px 0; border-radius: 10px; text-align: left; font-style: italic; color: #4c1d95;">
                            "{reason}"
                        </div>
                        <p style="font-size: 15px; color: #475569; line-height: 1.6; text-align: left;">
                            Please join us in congratulating <b>{name}</b> on this well-deserved recognition and wishing them continued success!
                        </p>
                        <hr style="border: 0; border-top: 1px solid #e2e8f0; margin: 25px 0;">
                        <p style="font-size: 12px; color: #94a3b8;">People Operations & HR Team</p>
                    </div>
                </body>
            </html>
            """
            
            broadcast_emails = []
            for rec in recipients:
                rec_company_email = rec.get("email")
                if rec_company_email:
                    broadcast_emails.append(rec_company_email)
            
            if broadcast_emails:
                send_email(", ".join(broadcast_emails), announce_subject, announce_body, inline_images=inline_images, server=smtp_server)
        
        db.appreciations.update_one({"_id": ObjectId(appreciation_id)}, {"$set": {"status": "SENT"}})
        if smtp_server:
            try:
                smtp_server.quit()
            except Exception:
                pass
        return jsonify({"message": "Appreciation and announcement emails dispatched successfully!"})
    except Exception as e:
        if smtp_server:
            try:
                smtp_server.quit()
            except Exception:
                pass
        return jsonify({"detail": str(e)}), 500

@app.route('/api/surprise-ops/appreciation/<appreciation_id>', methods=['DELETE'])
@login_required
def delete_appreciation(appreciation_id):
    try:
        tenant_id = g.current_user["tenant_id"]
        appr = db.appreciations.find_one({"_id": ObjectId(appreciation_id), "tenant_id": tenant_id})
        if appr and appr.get("certificate_url"):
            import os
            rel_path = appr["certificate_url"].lstrip('/')
            filepath = os.path.join(app.root_path, rel_path)
            if os.path.exists(filepath):
                try:
                    os.remove(filepath)
                except OSError:
                    pass
        res = db.appreciations.delete_one({"_id": ObjectId(appreciation_id), "tenant_id": tenant_id})
        if res.deleted_count == 0:
            return jsonify({"detail": "Record not found"}), 404
        return jsonify({"message": "Appreciation record deleted successfully."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- LMS & Club ---
@app.route('/api/lms-club/clubs', methods=['GET', 'POST'])
@login_required
def handle_clubs():
    try:
        tenant_id = g.current_user["tenant_id"]
        if request.method == 'POST':
            data = request.json or {}
            name = data.get('name')
            description = data.get('description')
            
            if not name:
                return jsonify({"detail": "name is required"}), 400
                
            res = db.clubs.insert_one({
                "tenant_id": tenant_id,
                "name": name,
                "description": description,
                "members": []
            })
            club = db.clubs.find_one({"_id": res.inserted_id})
            return jsonify(serialize_doc(club))
        else:
            clubs = list(db.clubs.find({"tenant_id": tenant_id}))
            result = []
            for c in clubs:
                member_ids = c.get("members", [])
                members = list(db.employees.find({"_id": {"$in": member_ids}}))
                result.append({
                    "id": str(c["_id"]),
                    "name": c["name"],
                    "description": c.get("description", ""),
                    "members": [{"id": str(m["_id"]), "name": m["name"], "email": m["email"]} for m in members]
                })
            return jsonify(result)
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/lms-club/clubs/<club_id>', methods=['DELETE'])
@login_required
def delete_club(club_id):
    try:
        db.clubs.delete_one({"_id": ObjectId(club_id)})
        return jsonify({"message": "Club deleted successfully"})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/lms-club/clubs/<club_id>/members', methods=['POST'])
@login_required
def update_club_members(club_id):
    data = request.json or {}
    employee_ids = data.get('employee_ids', [])
    
    try:
        # Convert string employee IDs to ObjectIds
        obj_ids = [ObjectId(eid) for eid in employee_ids]
        db.clubs.update_one({"_id": ObjectId(club_id)}, {"$set": {"members": obj_ids}})
        
        club = db.clubs.find_one({"_id": ObjectId(club_id)})
        members = list(db.employees.find({"_id": {"$in": obj_ids}}))
        
        return jsonify({
            "id": str(club["_id"]),
            "name": club["name"],
            "members": [{"id": str(m["_id"]), "name": m["name"], "email": m["email"]} for m in members]
        })
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- E-Library ---
import uuid as uuid_lib
class ElibraryDir:
    def __str__(self):
        return get_upload_dir('elibrary')
    def __fspath__(self):
        return get_upload_dir('elibrary')

ELIBRARY_UPLOAD_DIR = ElibraryDir()

@app.route('/api/elibrary', methods=['GET', 'POST'])
@login_required
def handle_elibrary():
    try:
        tenant_id = g.current_user["tenant_id"]
        role = g.current_user.get("role", "")

        if request.method == 'POST':
            # Only Admin (HR) can upload
            if role != 'Admin (HR)':
                return jsonify({"detail": "Only admins can upload files."}), 403
            if 'file' not in request.files:
                return jsonify({"detail": "No file provided."}), 400
            file = request.files['file']
            if file.filename == '':
                return jsonify({"detail": "Empty filename."}), 400
            ext = os.path.splitext(file.filename)[1].lower()
            if ext not in ['.pdf', '.xlsx', '.docx']:
                return jsonify({"detail": "Only .pdf, .xlsx, .docx files are allowed."}), 400
            os.makedirs(ELIBRARY_UPLOAD_DIR, exist_ok=True)
            from werkzeug.utils import secure_filename
            safe_name = secure_filename(file.filename)
            unique_name = f"{uuid_lib.uuid4().hex}_{safe_name}"
            filepath = os.path.join(ELIBRARY_UPLOAD_DIR, unique_name)
            file.save(filepath)
            title = request.form.get('title') or safe_name
            description = request.form.get('description', '')
            doc = {
                "tenant_id": tenant_id,
                "title": title,
                "description": description,
                "filename": unique_name,
                "original_name": file.filename,
                "ext": ext,
                "uploaded_by": g.current_user.get("name", "Admin"),
                "uploaded_at": datetime.datetime.utcnow().isoformat()
            }
            res = db.elibrary.insert_one(doc)
            doc["id"] = str(res.inserted_id)
            doc.pop("_id", None)
            return jsonify(doc), 201
        else:
            docs = list(db.elibrary.find({"tenant_id": tenant_id}))
            return jsonify([{
                "id": str(d["_id"]),
                "title": d.get("title", d.get("original_name", "")),
                "description": d.get("description", ""),
                "filename": d["filename"],
                "original_name": d.get("original_name", ""),
                "ext": d.get("ext", ""),
                "uploaded_by": d.get("uploaded_by", ""),
                "uploaded_at": d.get("uploaded_at", "")
            } for d in docs])
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/elibrary/<file_id>', methods=['DELETE'])
@login_required
def delete_elibrary_file(file_id):
    try:
        role = g.current_user.get("role", "")
        if role != 'Admin (HR)':
            return jsonify({"detail": "Only admins can delete files."}), 403
        tenant_id = g.current_user["tenant_id"]
        doc = db.elibrary.find_one({"_id": ObjectId(file_id), "tenant_id": tenant_id})
        if not doc:
            return jsonify({"detail": "File not found."}), 404
        filepath = os.path.join(ELIBRARY_UPLOAD_DIR, doc["filename"])
        if os.path.exists(filepath):
            os.remove(filepath)
        db.elibrary.delete_one({"_id": ObjectId(file_id)})
        return jsonify({"message": "File deleted successfully."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/static/uploads/elibrary/<path:filename>')
@login_required
def serve_elibrary_file(filename):
    return send_from_directory(ELIBRARY_UPLOAD_DIR, filename)

# --- E-Library External Links & Courses ---
@app.route('/api/elibrary/links', methods=['GET', 'POST'])
@login_required
def handle_elibrary_links():
    try:
        tenant_id = g.current_user["tenant_id"]
        role = g.current_user.get("role", "")

        if request.method == 'POST':
            # Only Admin (HR) can add links
            if role != 'Admin (HR)':
                return jsonify({"detail": "Only admins can add course links."}), 403
            data = request.json or {}
            title = data.get('title', '').strip()
            url = data.get('url', '').strip()
            description = data.get('description', '').strip()

            if not title or not url:
                return jsonify({"detail": "Title and URL are required."}), 400

            doc = {
                "tenant_id": tenant_id,
                "title": title,
                "url": url,
                "description": description,
                "uploaded_by": g.current_user.get("name", "Admin"),
                "uploaded_at": datetime.datetime.utcnow().isoformat()
            }
            res = db.elibrary_links.insert_one(doc)
            doc["id"] = str(res.inserted_id)
            doc.pop("_id", None)
            return jsonify(doc), 201
        else:
            docs = list(db.elibrary_links.find({"tenant_id": tenant_id}))
            return jsonify([{
                "id": str(d["_id"]),
                "title": d.get("title", ""),
                "url": d.get("url", ""),
                "description": d.get("description", ""),
                "uploaded_by": d.get("uploaded_by", ""),
                "uploaded_at": d.get("uploaded_at", "")
            } for d in docs])
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/elibrary/links/<link_id>', methods=['DELETE'])
@login_required
def delete_elibrary_link(link_id):
    try:
        role = g.current_user.get("role", "")
        if role != 'Admin (HR)':
            return jsonify({"detail": "Only admins can delete course links."}), 403
        tenant_id = g.current_user["tenant_id"]
        res = db.elibrary_links.delete_one({"_id": ObjectId(link_id), "tenant_id": tenant_id})
        if res.deleted_count == 0:
            return jsonify({"detail": "Link not found."}), 404
        return jsonify({"message": "Link deleted successfully."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500


# --- Discussions ---
@app.route('/api/discussions', methods=['GET', 'POST'])
@login_required
def handle_discussions():
    try:
        tenant_id = g.current_user["tenant_id"]
        role = g.current_user.get("role", "")
        emp_id = str(g.current_user.get("employee_id", ""))

        if request.method == 'POST':
            # Only Employees can create threads
            if role == 'Admin (HR)':
                return jsonify({"detail": "Admins cannot create discussion threads."}), 403
            data = request.json or {}
            title = data.get("title", "").strip()
            body = data.get("body", "").strip()
            venue = data.get("venue", "").strip()
            invited_ids = data.get("invited_ids", [])  # list of employee IDs
            if not title:
                return jsonify({"detail": "Thread title is required."}), 400
            if not venue:
                return jsonify({"detail": "Venue is required."}), 400

            # Fetch invited employees for email dispatch
            invited_employees = []
            if invited_ids:
                obj_ids = [ObjectId(eid) for eid in invited_ids if eid]
                invited_employees = list(db.employees.find({"_id": {"$in": obj_ids}}))

            thread = {
                "tenant_id": tenant_id,
                "title": title,
                "body": body,
                "venue": venue,
                "creator_id": emp_id,
                "creator_name": g.current_user.get("name", "Employee"),
                "participants": [emp_id] + [str(e["_id"]) for e in invited_employees],
                "messages": [],
                "created_at": datetime.datetime.utcnow().isoformat()
            }
            res = db.discussions.insert_one(thread)
            thread_id = str(res.inserted_id)

            # Send email invitations to each invited employee
            creator_name = g.current_user.get("name", "A colleague")
            for emp in invited_employees:
                emails = []
                if emp.get("email"): emails.append(emp["email"])
                if emp.get("personal_email"): emails.append(emp["personal_email"])
                to_email = ", ".join(e for e in emails if e)
                if not to_email:
                    continue
                invite_body = f"""
                <html><body style="font-family:Arial,sans-serif;background:#f8fafc;padding:20px;color:#1e293b;">
                  <div style="background:white;padding:30px;border-radius:12px;max-width:600px;margin:0 auto;">
                    <h2 style="color:#f59e0b;border-bottom:2px solid #fef3c7;padding-bottom:10px;margin-top:0;">
                      💬 Discussion Thread Invitation
                    </h2>
                    <p style="font-size:16px;"><strong>{emp.get('name','Team Member')}</strong>,</p>
                    <p><strong>{creator_name}</strong> has invited you to join a discussion:</p>
                    <h3 style="color:#0f172a;">{title}</h3>
                    <p><strong>Venue:</strong> {venue}</p>
                    {f'<p style="color:#64748b;">{body}</p>' if body else ''}
                    <p style="font-size:13px;color:#94a3b8;margin-top:20px;">
                      Log in to the SEMCO Groups HR portal to view and participate in this thread.
                    </p>
                  </div>
                </body></html>
                """
                try:
                    send_email(to_email, f"Discussion Invite: {title}", invite_body)
                except Exception:
                    pass

            return jsonify({"id": thread_id, "title": title, "message": "Thread created and invites sent."}), 201
        else:
            # Admins see all threads; employees see threads where they are participants
            if role == 'Admin (HR)':
                threads = list(db.discussions.find({"tenant_id": tenant_id}))
            else:
                threads = list(db.discussions.find({"tenant_id": tenant_id, "participants": emp_id}))

            result = []
            for t in threads:
                # Lookup participant details
                p_ids = t.get("participants", [])
                p_obj_ids = []
                for pid in p_ids:
                    try:
                        p_obj_ids.append(ObjectId(pid))
                    except:
                        pass
                p_users = list(db.employees.find({"_id": {"$in": p_obj_ids}}))
                p_details = [{"id": str(u["_id"]), "name": u["name"], "email": u.get("email", "")} for u in p_users]

                result.append({
                    "id": str(t["_id"]),
                    "title": t.get("title", ""),
                    "body": t.get("body", ""),
                    "venue": t.get("venue", ""),
                    "creator_id": t.get("creator_id", ""),
                    "creator_name": t.get("creator_name", ""),
                    "participants": t.get("participants", []),
                    "participants_details": p_details,
                    "created_at": t.get("created_at", "")
                })

            return jsonify(result)
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/discussions/<thread_id>/messages', methods=['POST'])
@login_required
def post_discussion_message(thread_id):
    try:
        tenant_id = g.current_user["tenant_id"]
        role = g.current_user.get("role", "")
        if role == 'Admin (HR)':
            return jsonify({"detail": "Admins cannot post messages in discussions."}), 403
        data = request.json or {}
        text = data.get("text", "").strip()
        if not text:
            return jsonify({"detail": "Message text is required."}), 400
        message = {
            "sender_id": str(g.current_user.get("_id", "")),
            "sender_name": g.current_user.get("name", "Employee"),
            "text": text,
            "sent_at": datetime.datetime.utcnow().isoformat()
        }
        res = db.discussions.update_one(
            {"_id": ObjectId(thread_id), "tenant_id": tenant_id},
            {"$push": {"messages": message}}
        )
        if res.matched_count == 0:
            return jsonify({"detail": "Thread not found."}), 404
        return jsonify({"message": "Message posted."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/discussions/<thread_id>', methods=['DELETE'])
@login_required
def delete_discussion(thread_id):
    try:
        tenant_id = g.current_user["tenant_id"]
        role = g.current_user.get("role", "")
        emp_id = str(g.current_user.get("_id", ""))
        thread = db.discussions.find_one({"_id": ObjectId(thread_id), "tenant_id": tenant_id})
        if not thread:
            return jsonify({"detail": "Thread not found."}), 404
        # Only the creator can delete their own thread
        if role != 'Admin (HR)' and thread.get("creator_id") != emp_id:
            return jsonify({"detail": "Only the thread creator can delete this thread."}), 403
        db.discussions.delete_one({"_id": ObjectId(thread_id)})
        return jsonify({"message": "Thread deleted."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- SSE Event Subscribers ---
import queue as queue_module
_event_subscribers = []

def _broadcast_event_update(payload: str):
    """Notify all active SSE subscribers of an event change."""
    dead = []
    for q in list(_event_subscribers):
        try:
            q.put_nowait(payload)
        except Exception:
            dead.append(q)
    for q in dead:
        try:
            _event_subscribers.remove(q)
        except ValueError:
            pass

# --- Events Planner ---
@app.route('/api/events', methods=['GET', 'POST'])
@login_required
def handle_events():
    try:
        tenant_id = g.current_user["tenant_id"]
        if request.method == 'POST':
            data = request.json or {}
            title = data.get('title')
            description = data.get('description')
            start_time_str = data.get('start_time')
            location = data.get('location')
            
            attendees = data.get('attendees', [])
            
            if not title or not start_time_str:
                return jsonify({"detail": "title and start_time are required"}), 400
                
            res = db.events.insert_one({
                "tenant_id": tenant_id,
                "title": title,
                "description": description,
                "start_time": start_time_str,
                "location": location,
                "attendees": attendees,
                "status": "ACTIVE"
            })
            event = db.events.find_one({"_id": res.inserted_id})
            
            # Queue APScheduler notifications
            start_time_dt = datetime.datetime.fromisoformat(start_time_str.replace('Z', ''))
            queue_event_reminders(str(event["_id"]), event["title"], start_time_dt)
            
            # Broadcast to SSE subscribers
            import json as _json
            _broadcast_event_update(_json.dumps({"action": "created", "id": str(event["_id"])}))
            
            return jsonify({
                "id": str(event["_id"]),
                "title": event["title"],
                "description": event["description"],
                "start_time": event["start_time"],
                "location": event["location"],
                "attendees": event.get("attendees", []),
                "status": event.get("status", "ACTIVE"),
                "is_exit_interview": event.get("is_exit_interview", False),
                "employee_id": event.get("employee_id", "")
            })
        else:
            # Auto-archive past completed events
            now_iso = datetime.datetime.now().isoformat()
            db.events.update_many(
                {"tenant_id": tenant_id, "start_time": {"$lt": now_iso}, "status": {"$ne": "ARCHIVED"}},
                {"$set": {"status": "ARCHIVED"}}
            )
            
            events = list(db.events.find({"tenant_id": tenant_id}))
            
            # Filter events for employee views (non-Admins)
            role = g.current_user.get("role", "")
            if role != 'Admin (HR)':
                emp_id_str = str(g.current_user.get("employee_id", ""))
                filtered_events = []
                for ev in events:
                    is_exit = ev.get("is_exit_interview") or "exit interview" in ev.get("title", "").lower()
                    if is_exit:
                        # Exit interview is only visible to the specific employee being offboarded
                        if ev.get("employee_id") == emp_id_str:
                            filtered_events.append(ev)
                        elif not ev.get("employee_id"):
                            # Fallback: check if employee name matches in title (e.g. 'Exit Interview: Mrunal')
                            title_lower = ev.get("title", "").lower()
                            emp_name_lower = g.current_user.get("name", "").lower()
                            if emp_name_lower and emp_name_lower in title_lower:
                                filtered_events.append(ev)
                    else:
                        # Regular events are visible if attendee list is empty (public) or if the employee is an attendee
                        attendees = ev.get("attendees", [])
                        if not attendees or any(str(att.get("id")) == emp_id_str for att in attendees):
                            filtered_events.append(ev)
                events = filtered_events
                
            return jsonify([{
                "id": str(ev["_id"]),
                "title": ev["title"],
                "description": ev.get("description", ""),
                "start_time": ev["start_time"],
                "location": ev.get("location", ""),
                "attendees": ev.get("attendees", []),
                "status": ev.get("status", "ACTIVE"),
                "is_exit_interview": ev.get("is_exit_interview", False),
                "employee_id": ev.get("employee_id", "")
            } for ev in events])
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/events/<event_id>/archive', methods=['PUT'])
@login_required
def archive_event(event_id):
    try:
        tenant_id = g.current_user["tenant_id"]
        res = db.events.update_one(
            {"_id": ObjectId(event_id), "tenant_id": tenant_id},
            {"$set": {"status": "ARCHIVED"}}
        )
        if res.matched_count == 0:
            return jsonify({"detail": "Event not found"}), 404
        import json as _json
        _broadcast_event_update(_json.dumps({"action": "archived", "id": event_id}))
        return jsonify({"message": "Event archived successfully."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/events/<event_id>', methods=['DELETE'])
@login_required
def delete_event(event_id):
    try:
        tenant_id = g.current_user["tenant_id"]
        
        # Remove any queued APScheduler reminder jobs for this event
        from app.utils.scheduler import scheduler
        intervals = ["30_days_before", "7_days_before", "1_day_before", "30_minutes_before"]
        for name in intervals:
            job_id = f"event_{event_id}_{name}"
            try:
                if scheduler.get_job(job_id):
                    scheduler.remove_job(job_id)
            except Exception:
                pass
                
        res = db.events.delete_one({"_id": ObjectId(event_id), "tenant_id": tenant_id})
        if res.deleted_count == 0:
            return jsonify({"detail": "Event not found"}), 404
        import json as _json
        _broadcast_event_update(_json.dumps({"action": "deleted", "id": event_id}))
        return jsonify({"message": "Event deleted forever."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/events/stream')
@login_required
def events_stream():
    """Server-Sent Events endpoint: streams event change notifications to connected clients."""
    import json as _json
    from flask import Response, stream_with_context
    q = queue_module.Queue(maxsize=50)
    _event_subscribers.append(q)
    def generate():
        try:
            yield f"data: {_json.dumps({'action': 'connected'})}\n\n"
            while True:
                try:
                    payload = q.get(timeout=25)
                    yield f"data: {payload}\n\n"
                except queue_module.Empty:
                    yield "data: {\"action\":\"ping\"}\n\n"
        except GeneratorExit:
            pass
        finally:
            try:
                _event_subscribers.remove(q)
            except ValueError:
                pass
    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection': 'keep-alive'
        }
    )

@app.route('/api/holidays', methods=['GET'])
@login_required
def get_holidays():
    try:
        year_param = request.args.get('year', type=int)
        if year_param:
            # Dynamically seed the year if it doesn't have any holiday data yet
            from app.utils.scheduler import seed_holidays_for_year
            seed_holidays_for_year(year_param)
            
        holidays = list(db.holidays.find({}))
        return jsonify([{
            "id": str(h["_id"]),
            "name": h["name"],
            "date": h["date"],
            "type": h.get("type", "National")
        } for h in holidays])
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/holidays', methods=['POST'])
@login_required
def create_holiday():
    try:
        tenant_id = g.current_user["tenant_id"]
        data = request.json or {}
        name = data.get("name")
        date = data.get("date")
        h_type = data.get("type", "Bank")
        if not name or not date:
            return jsonify({"detail": "name and date are required"}), 400
            
        db.holidays.insert_one({
            "name": name,
            "date": date,
            "type": h_type
        })
        
        # Dispatch notification emails to all registered active employees
        employees = list(db.employees.find({"tenant_id": tenant_id, "status": "ACTIVE"}))
        recipient_emails = []
        for emp in employees:
            email = emp.get("email")
            if email:
                recipient_emails.append(email)
                
        if recipient_emails:
            try:
                parsed_date = datetime.datetime.strptime(date, "%Y-%m-%d")
                formatted_date = parsed_date.strftime("%A, %B %d, %Y")
            except Exception:
                formatted_date = date
                
            subject = f"Notice: Upcoming Holiday on {date} - {name} 📅🎉"
            body = f"""
            <html>
                <body style="font-family: Arial, sans-serif; background-color: #faf5ff; padding: 20px; color: #1e293b;">
                    <div style="background-color: white; padding: 40px; border-radius: 16px; max-width: 600px; margin: 0 auto; box-shadow: 0 10px 15px -3px rgba(0,0,0,0.1); border-top: 8px solid #14b8a6; text-align: center;">
                        <span style="font-size: 50px;">📅🎉</span>
                        <h2 style="color: #0d9488; margin-top: 15px; font-size: 22px;">Holiday Announcement</h2>
                        <p style="font-size: 15px; color: #475569; line-height: 1.6; text-align: left; margin-top: 20px;">
                            Dear Team,<br><br>
                            Please be informed that the company will observe a holiday on <b>{formatted_date}</b> on account of <b>{name}</b> ({h_type} Holiday).
                        </p>
                        <div style="background-color: #f0fdfa; border: 1px solid #99f6e4; padding: 20px; margin: 20px 0; border-radius: 10px; text-align: center; color: #0f766e; font-weight: bold; font-size: 16px;">
                            ✨ Enjoy the upcoming holiday! ✨
                        </div>
                        <p style="font-size: 15px; color: #475569; line-height: 1.6; text-align: left;">
                            We hope you have a relaxing day off. Please make sure to wrap up your critical tasks before the holiday.
                        </p>
                        <hr style="border: 0; border-top: 1px solid #e2e8f0; margin: 25px 0;">
                        <p style="font-size: 12px; color: #94a3b8;">People Operations & HR Team</p>
                    </div>
                </body>
            </html>
            """
            send_email(", ".join(recipient_emails), subject, body)
            
        return jsonify({"message": "Holiday added successfully."}), 201
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/holidays/<holiday_id>', methods=['PUT', 'DELETE'])
@login_required
def update_or_delete_holiday(holiday_id):
    try:
        if request.method == 'DELETE':
            res = db.holidays.delete_one({"_id": ObjectId(holiday_id)})
            if res.deleted_count == 0:
                return jsonify({"detail": "Holiday not found"}), 404
            return jsonify({"message": "Holiday deleted successfully."})
            
        data = request.json or {}
        update_fields = {}
        if "name" in data:
            update_fields["name"] = data["name"]
        if "date" in data:
            update_fields["date"] = data["date"]
        if "type" in data:
            update_fields["type"] = data["type"]
            
        res = db.holidays.update_one({"_id": ObjectId(holiday_id)}, {"$set": update_fields})
        if res.matched_count == 0:
            return jsonify({"detail": "Holiday not found"}), 404
        return jsonify({"message": "Holiday updated successfully."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Asset Management Portal ---
@app.route('/api/assets', methods=['GET', 'POST'])
@login_required
def handle_assets():
    try:
        tenant_id = g.current_user["tenant_id"]
        if request.method == 'POST':
            data = request.json or {}
            hardware_name = data.get('hardware_name')
            serial_number = data.get('serial_number')
            category = data.get('category', 'General')
            try:
                total_quantity = int(data.get('total_quantity', 1))
            except Exception:
                total_quantity = 1
                
            if not hardware_name or not serial_number:
                return jsonify({"detail": "hardware_name and serial_number are required"}), 400
                
            res = db.assets.insert_one({
                "tenant_id": tenant_id,
                "hardware_name": hardware_name,
                "serial_number": serial_number,
                "category": category,
                "total_quantity": total_quantity,
                "status": "AVAILABLE",
                "assigned_to": None,
                "checkout_date": None,
                "due_date": None
            })
            asset = db.assets.find_one({"_id": res.inserted_id})
            return jsonify(serialize_doc(asset))
        else:
            assets = list(db.assets.find({"tenant_id": tenant_id}))
            
            # Gather all employee IDs from all checkouts plus the legacy assigned_to
            emp_ids = []
            for a in assets:
                if a.get("assigned_to"):
                    emp_ids.append(ObjectId(a["assigned_to"]))
                for c in a.get("checkouts") or []:
                    if c.get("employee_id"):
                        emp_ids.append(ObjectId(c["employee_id"]))
            
            employees = list(db.employees.find({"_id": {"$in": emp_ids}}))
            emp_map = {str(e["_id"]): e for e in employees}
            
            result = []
            for a in assets:
                checkouts_list = []
                total_checked_out = 0
                for c in a.get("checkouts") or []:
                    emp_info = emp_map.get(str(c["employee_id"]), {})
                    qty = c.get("quantity", 1)
                    total_checked_out += qty
                    checkouts_list.append({
                        "id": c.get("id"),
                        "employee_id": str(c["employee_id"]),
                        "employee_name": emp_info.get("name", "Unknown"),
                        "employee_email": emp_info.get("email", ""),
                        "quantity": qty,
                        "checkout_date": c.get("checkout_date"),
                        "due_date": c.get("due_date")
                    })
                
                total_qty = a.get("total_quantity", 1)
                remaining_qty = max(0, total_qty - total_checked_out)
                status = "CHECKED_OUT" if remaining_qty <= 0 else "AVAILABLE"
                
                # Fetch last assigned employee for backwards compatibility
                legacy_emp = emp_map.get(str(a["assigned_to"]), {}) if a.get("assigned_to") else None
                
                result.append({
                    "id": str(a["_id"]),
                    "hardware_name": a.get("hardware_name", ""),
                    "serial_number": a.get("serial_number", ""),
                    "category": a.get("category", "General"),
                    "total_quantity": total_qty,
                    "remaining_quantity": remaining_qty,
                    "status": status,
                    "assigned_to": str(a["assigned_to"]) if a.get("assigned_to") else None,
                    "assigned_employee": {"name": legacy_emp.get("name", ""), "email": legacy_emp.get("email", "")} if legacy_emp else None,
                    "checkout_date": a.get("checkout_date"),
                    "due_date": a.get("due_date"),
                    "checkouts": checkouts_list
                })
            return jsonify(result)
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/assets/<asset_id>/checkout', methods=['POST'])
@login_required
def checkout_asset(asset_id):
    import uuid
    data = request.json or {}
    employee_id = data.get('employee_id')
    duration_days = int(data.get('duration_days', 14))
    try:
        quantity = int(data.get('quantity', 1))
    except Exception:
        quantity = 1
    
    if not employee_id:
        return jsonify({"detail": "employee_id is required"}), 400
        
    try:
        asset = db.assets.find_one({"_id": ObjectId(asset_id)})
        emp = db.employees.find_one({"_id": ObjectId(employee_id)})
        
        if not asset or not emp:
            return jsonify({"detail": "Asset or Employee not found"}), 404
            
        checkouts = asset.get("checkouts") or []
        total_checked_out = sum(c.get("quantity", 1) for c in checkouts)
        total_qty = asset.get("total_quantity", 1)
        remaining_qty = max(0, total_qty - total_checked_out)
        
        if quantity > remaining_qty:
            return jsonify({"detail": f"Only {remaining_qty} units available. Cannot checkout {quantity}."}), 400
            
        checkout_date = datetime.datetime.now().isoformat()
        due_date = (datetime.datetime.now() + datetime.timedelta(days=duration_days)).isoformat()
        
        new_checkout = {
            "id": str(uuid.uuid4()),
            "employee_id": ObjectId(employee_id),
            "quantity": quantity,
            "checkout_date": checkout_date,
            "due_date": due_date
        }
        
        new_remaining_qty = remaining_qty - quantity
        new_status = "CHECKED_OUT" if new_remaining_qty <= 0 else "AVAILABLE"
        
        db.assets.update_one({"_id": ObjectId(asset_id)}, {
            "$push": {"checkouts": new_checkout},
            "$set": {
                "assigned_to": ObjectId(employee_id),
                "status": new_status,
                "checkout_date": checkout_date,
                "due_date": due_date
            }
        })
        
        subject = f"Hardware Checkout Confirmation: {asset['hardware_name']}"
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; background-color: #f8fafc; padding: 20px; color: #1e293b;">
                <div style="background-color: white; padding: 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">
                    <h2 style="color: #06b6d4; border-bottom: 2px solid #ecfeff; padding-bottom: 10px; margin-top: 0;">Hardware Checked Out 💻</h2>
                    <p style="font-size: 15px; color: #475569;">
                        Dear {emp['name']},<br><br>
                        This is to confirm that you have checked out company hardware:
                    </p>
                    <table style="width:100%; border-collapse:collapse; margin:20px 0;">
                        <tr><td style="padding:6px; color:#64748b;"><b>Item:</b></td><td style="padding:6px;">{asset['hardware_name']}</td></tr>
                        <tr><td style="padding:6px; color:#64748b;"><b>Serial Number:</b></td><td style="padding:6px;">{asset['serial_number']}</td></tr>
                        <tr><td style="padding:6px; color:#64748b;"><b>Quantity Checked Out:</b></td><td style="padding:6px;"><b>{quantity}</b></td></tr>
                    </table>
                    <p style="font-size: 13px; color: #94a3b8;">Please keep this device secure.</p>
                </div>
            </body>
        </html>
        """
        send_email(emp["email"], subject, body)
        return jsonify({"message": f"Asset successfully checked out!"})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/assets/<asset_id>/checkin', methods=['POST'])
@login_required
def checkin_asset(asset_id):
    data = request.json or {}
    checkout_id = data.get('checkout_id')
    
    try:
        asset = db.assets.find_one({"_id": ObjectId(asset_id)})
        if not asset:
            return jsonify({"detail": "Asset not found"}), 404
            
        checkouts = asset.get("checkouts") or []
        emp = None
        if checkout_id:
            checkout = next((c for c in checkouts if c.get("id") == checkout_id), None)
            if not checkout:
                return jsonify({"detail": "Checkout log not found"}), 404
            emp = db.employees.find_one({"_id": checkout["employee_id"]})
            db.assets.update_one({"_id": ObjectId(asset_id)}, {
                "$pull": {"checkouts": {"id": checkout_id}}
            })
        else:
            if checkouts:
                checkout = checkouts[0]
                emp = db.employees.find_one({"_id": checkout["employee_id"]})
                db.assets.update_one({"_id": ObjectId(asset_id)}, {
                    "$pop": {"checkouts": -1}
                })
        
        updated_asset = db.assets.find_one({"_id": ObjectId(asset_id)})
        new_checkouts = updated_asset.get("checkouts") or []
        total_checked_out = sum(c.get("quantity", 1) for c in new_checkouts)
        total_qty = updated_asset.get("total_quantity", 1)
        new_remaining_qty = max(0, total_qty - total_checked_out)
        
        new_status = "AVAILABLE" if new_remaining_qty > 0 else "CHECKED_OUT"
        last_emp_id = new_checkouts[-1]["employee_id"] if new_checkouts else None
        
        db.assets.update_one({"_id": ObjectId(asset_id)}, {"$set": {
            "status": new_status,
            "assigned_to": last_emp_id
        }})
        
        if emp:
            subject = f"Hardware Return Confirmation: {asset['hardware_name']}"
            body = f"""
            <html>
                <body style="font-family: Arial, sans-serif; background-color: #f8fafc; padding: 20px; color: #1e293b;">
                    <div style="background-color: white; padding: 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">
                        <h2 style="color: #10b981; border-bottom: 2px solid #ecfdf5; padding-bottom: 10px; margin-top: 0;">Hardware Returned 💻✅</h2>
                        <p style="font-size: 15px; color: #475569;">
                            Dear {emp['name']},<br><br>
                            This is to confirm that you have successfully returned <b>{asset['hardware_name']}</b> (S/N: {asset['serial_number']}) back to HR Inventory.
                        </p>
                        <p style="font-size: 13px; color: #94a3b8;">Thank you for coordinating with HR Operations.</p>
                    </div>
                </body>
            </html>
            """
            send_email(emp["email"], subject, body)
            
        return jsonify({"message": "Device returned to inventory successfully."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Document Vault (Self-Service Certificates) ---
@app.route('/api/documents/requests', methods=['GET', 'POST'])
@login_required
def handle_documents():
    try:
        if request.method == 'POST':
            data = request.json or {}
            employee_id = data.get('employee_id')
            doc_type = data.get('document_type', 'EMPLOYMENT_PROOF')
            
            if not employee_id:
                return jsonify({"detail": "employee_id is required"}), 400
                
            emp = db.employees.find_one({"_id": ObjectId(employee_id)})
            if not emp:
                return jsonify({"detail": "Employee not found"}), 404
                
            request_date = datetime.datetime.now().isoformat()
            res = db.document_requests.insert_one({
                "employee_id": ObjectId(employee_id),
                "document_type": doc_type,
                "status": "PENDING",
                "request_date": request_date
            })
            req_id = res.inserted_id
            
            output_pdf = f"cert_employment_{str(req_id)}.pdf"
            
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            
            doc = SimpleDocTemplate(output_pdf, pagesize=letter, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
            story = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle('CertTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=22, spaceAfter=20, alignment=1)
            body_style = ParagraphStyle('CertBody', parent=styles['Normal'], fontName='Helvetica', fontSize=12, leading=18, spaceAfter=15)
            
            story.append(Paragraph("<b>TO WHOMSOEVER IT MAY CONCERN</b>", title_style))
            story.append(Spacer(1, 20))
            
            date_str = datetime.date.today().strftime('%B %d, %Y')
            story.append(Paragraph(f"<b>Date:</b> {date_str}", body_style))
            story.append(Spacer(1, 10))
            
            joining_date_formatted = ""
            if emp.get("joining_date"):
                try:
                    joining_date_formatted = datetime.date.fromisoformat(emp["joining_date"][:10]).strftime('%B %d, %Y')
                except Exception:
                    joining_date_formatted = emp["joining_date"]
            
            cert_body = f"""
            This is to certify that <b>{emp["name"]}</b> is an employee of <b>Acme Corporation</b>.
            They have been employed with us since <b>{joining_date_formatted}</b> 
            and currently hold the position of <b>{emp["role"]}</b> in the <b>{emp["department"]}</b> department.
            Their employment status is currently active and in good standing.
            <br><br>
            This document is issued upon the employee's request as proof of employment.
            """
            story.append(Paragraph(cert_body, body_style))
            story.append(Spacer(1, 40))
            
            story.append(Paragraph("Sincerely,<br><b>People Operations Director</b><br>Acme Corporation", body_style))
            doc.build(story)
            
            subject = f"Your Requested Document: {doc_type.replace('_', ' ').title()}"
            body = f"""
            <html>
                <body style="font-family: Arial, sans-serif; padding: 20px; color: #1e293b;">
                    <div style="background-color: white; padding: 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">
                        <h2 style="color: #14b8a6; border-bottom: 2px solid #f0fdfa; padding-bottom: 10px; margin-top: 0;">Self-Service Document Vault 📑</h2>
                        <p style="font-size: 15px; color: #475569;">
                            Dear {emp["name"]},<br><br>
                            Your requested employment certification document has been automatically generated using our verified template and is attached to this email.
                        </p>
                        <p style="font-size: 13px; color: #94a3b8;">This is an automated request execution. No signature is physically required.</p>
                    </div>
                </body>
            </html>
            """
            
            send_email(emp["email"], subject, body, attachment_path=output_pdf, attachment_name=f"Employment_Proof_{emp['name'].replace(' ', '_')}.pdf")
            
            if os.path.exists(output_pdf):
                try:
                    os.remove(output_pdf)
                except OSError:
                    pass
                    
            db.document_requests.update_one({"_id": req_id}, {"$set": {"status": "GENERATED"}})
            
            return jsonify({
                "id": str(req_id),
                "employee_id": employee_id,
                "document_type": doc_type,
                "status": "GENERATED",
                "request_date": request_date
            })
        else:
            tenant_id = g.current_user["tenant_id"]
            employees = list(db.employees.find({"tenant_id": tenant_id}))
            emp_ids = [e["_id"] for e in employees]
            emp_map = {str(e["_id"]): e for e in employees}
            
            requests = list(db.document_requests.find({"employee_id": {"$in": emp_ids}}))
            
            result = []
            for r in requests:
                emp_info = emp_map.get(str(r["employee_id"]), {})
                result.append({
                    "id": str(r["_id"]),
                    "employee_id": str(r["employee_id"]),
                    "document_type": r["document_type"],
                    "status": r["status"],
                    "request_date": r["request_date"],
                    "employee": {
                        "name": emp_info.get("name", ""),
                        "email": emp_info.get("email", "")
                    }
                })
            return jsonify(result)
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Document Vault (Advanced Employee Self-Service & Admin HR) ---
@app.route('/api/documents/vault/<employee_id>', methods=['GET'])
@login_required
def get_document_vault(employee_id):
    try:
        tenant_id = g.current_user["tenant_id"]
        if employee_id == 'me':
            target_emp_id = ObjectId(g.current_user["employee_id"])
        else:
            target_emp_id = ObjectId(employee_id)
            
        doc = db.employee_documents.find_one({"employee_id": target_emp_id, "tenant_id": tenant_id})
        if not doc:
            return jsonify({
                "employee_id": str(target_emp_id),
                "personal_documents": {},
                "company_documents": []
            })
            
        res_doc = serialize_doc(doc)
        if res_doc and "employee_id" in res_doc:
            res_doc["employee_id"] = str(res_doc["employee_id"])
        return jsonify(res_doc)
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/documents/personal', methods=['POST'])
@login_required
def upload_personal_documents():
    try:
        tenant_id = g.current_user["tenant_id"]
        employee_id = ObjectId(g.current_user["employee_id"])
        
        # 1. Enforce strict size limit of 10MB per file
        MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
        for key in request.files:
            file = request.files[key]
            if file and file.filename != '':
                file.seek(0, os.SEEK_END)
                size = file.tell()
                file.seek(0)
                if size > MAX_FILE_SIZE:
                    return jsonify({"detail": f"File '{file.filename}' exceeds the strict 10MB limit."}), 400

        # Check existing documents
        existing_doc = db.employee_documents.find_one({"employee_id": employee_id, "tenant_id": tenant_id})
        existing_personal = existing_doc.get("personal_documents", {}) if existing_doc else {}
        
        # 2. Extract and validate text inputs
        bank_name = request.form.get('bank_name')
        account_number = request.form.get('account_number')
        account_name = request.form.get('account_name')
        ifsc_code = request.form.get('ifsc_code')
        
        try:
            percentage_10th = float(request.form.get('percentage_10th', 0.0))
            percentage_12th = float(request.form.get('percentage_12th', 0.0))
            bachelors_cgpa = float(request.form.get('bachelors_cgpa', 0.0))
        except ValueError:
            return jsonify({"detail": "Academic performance metrics must be numeric."}), 400
            
        has_experience = request.form.get('has_experience') == 'true'
        
        if not bank_name or not account_number or not account_name or not ifsc_code:
            return jsonify({"detail": "All bank details are compulsory."}), 400
            
        if not percentage_10th or not percentage_12th or not bachelors_cgpa:
            return jsonify({"detail": "All educational percentages/CGPA are compulsory."}), 400

        # 3. File validation and upload handling
        import uuid
        from werkzeug.utils import secure_filename
        
        upload_dir = get_upload_dir('documents')
        
        def save_file(file_key, allowed_extensions=None, compulsory=False, existing_url=None):
            file = request.files.get(file_key)
            if not file or file.filename == '':
                if compulsory and not existing_url:
                    raise ValueError(f"Document '{file_key}' is compulsory.")
                return existing_url
                
            filename = file.filename.lower()
            if allowed_extensions:
                ext = os.path.splitext(filename)[1]
                if ext not in allowed_extensions:
                    raise ValueError(f"File format for '{file_key}' is restricted. Allowed: {', '.join(allowed_extensions)}")
                    
            b64_uri = convert_file_to_base64_uri(file, allowed_extensions)
            if not b64_uri:
                if compulsory and not existing_url:
                    raise ValueError(f"Document '{file_key}' is compulsory.")
                return existing_url
            return b64_uri

        try:
            aadhaar_url = save_file('aadhaar', allowed_extensions=['.pdf'], compulsory=True, existing_url=existing_personal.get('aadhaar_url'))
            pan_url = save_file('pan', allowed_extensions=['.pdf'], compulsory=True, existing_url=existing_personal.get('pan_url'))
            profile_photo_url = save_file('profile_photo', allowed_extensions=['.png', '.jpg', '.jpeg'], compulsory=True, existing_url=existing_personal.get('profile_photo_url'))
            academic_proofs_url = save_file('academic_proofs', compulsory=True, existing_url=existing_personal.get('academic_proofs_url'))
            
            passbook_url = save_file('passbook', compulsory=False, existing_url=existing_personal.get('passbook_url'))
            experience_url = save_file('experience', compulsory=False, existing_url=existing_personal.get('experience_url'))
        except ValueError as val_err:
            return jsonify({"detail": str(val_err)}), 400
            
        personal_data = {
            "aadhaar_url": aadhaar_url,
            "pan_url": pan_url,
            "bank_name": bank_name,
            "account_number": account_number,
            "account_name": account_name,
            "ifsc_code": ifsc_code,
            "passbook_url": passbook_url,
            "profile_photo_url": profile_photo_url,
            "percentage_10th": percentage_10th,
            "percentage_12th": percentage_12th,
            "bachelors_cgpa": bachelors_cgpa,
            "academic_proofs_url": academic_proofs_url,
            "has_experience": has_experience,
            "experience_url": experience_url
        }
        
        db.employee_documents.update_one(
            {"employee_id": employee_id, "tenant_id": tenant_id},
            {"$set": {
                "personal_documents": personal_data
            }},
            upsert=True
        )
        
        updated_doc = db.employee_documents.find_one({"employee_id": employee_id, "tenant_id": tenant_id})
        res_doc = serialize_doc(updated_doc)
        if res_doc and "employee_id" in res_doc:
            res_doc["employee_id"] = str(res_doc["employee_id"])
        return jsonify(res_doc)
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/documents/company/<employee_id>', methods=['POST'])
@login_required
def upload_company_documents(employee_id):
    try:
        tenant_id = g.current_user["tenant_id"]
        target_emp_id = ObjectId(employee_id)
        
        # 1. Enforce strict size limit of 10MB per file
        MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
        for key in request.files:
            file = request.files[key]
            if file and file.filename != '':
                file.seek(0, os.SEEK_END)
                size = file.tell()
                file.seek(0)
                if size > MAX_FILE_SIZE:
                    return jsonify({"detail": f"File '{file.filename}' exceeds the strict 10MB limit."}), 400

        import json
        existing_docs_json = request.form.get('existing_docs', '[]')
        try:
            existing_docs = json.loads(existing_docs_json)
        except Exception:
            existing_docs = []
            
        uploaded_files = request.files.getlist('files')
        titles = request.form.getlist('titles')
        
        import uuid
        from werkzeug.utils import secure_filename
        
        upload_dir = get_upload_dir('documents')
        
        new_docs = []
        for file, title in zip(uploaded_files, titles):
            if file and file.filename != '' and title:
                b64_uri = convert_file_to_base64_uri(file)
                if b64_uri:
                    new_docs.append({
                        "id": uuid.uuid4().hex,
                        "title": title,
                        "file_url": b64_uri,
                        "upload_date": datetime.datetime.now().isoformat()
                    })
                
        merged_docs = existing_docs + new_docs
        
        db.employee_documents.update_one(
            {"employee_id": target_emp_id, "tenant_id": tenant_id},
            {"$set": {
                "company_documents": merged_docs
            }},
            upsert=True
        )
        
        updated_doc = db.employee_documents.find_one({"employee_id": target_emp_id, "tenant_id": tenant_id})
        res_doc = serialize_doc(updated_doc)
        if res_doc and "employee_id" in res_doc:
            res_doc["employee_id"] = str(res_doc["employee_id"])
        return jsonify(res_doc)
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Offboarding Automator ---
@app.route('/api/offboarding/status', methods=['GET'])
@login_required
def get_my_offboarding_status():
    try:
        user_id = g.current_user["employee_id"]
        emp = db.employees.find_one({"_id": ObjectId(user_id)})
        if not emp:
            return jsonify({"detail": "Employee not found"}), 404
            
        return jsonify({
            "status": emp.get("status", "ACTIVE"),
            "notice_period_served": emp.get("notice_period_served"),
            "notice_period_duration": emp.get("notice_period_duration"),
            "notice_period_unit": emp.get("notice_period_unit"),
            "exit_interview_date": emp.get("exit_interview_date"),
            "exit_interview_time": emp.get("exit_interview_time"),
            "exit_interview_location": emp.get("exit_interview_location")
        })
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/offboarding/trigger/<employee_id>', methods=['POST'])
@login_required
def trigger_offboarding(employee_id):
    try:
        emp = db.employees.find_one({"_id": ObjectId(employee_id)})
        if not emp:
            return jsonify({"detail": "Employee not found"}), 404
            
        # 1. Update Employee Status
        db.employees.update_one({"_id": ObjectId(employee_id)}, {"$set": {
            "status": "Offboarding Initialized"
        }})
        
        # 2. Email Employee with portal link
        emp_subject = "Offboarding Initiated"
        target_email = emp.get("personal_email") or emp.get("email")
        scheme = "https" if request.is_secure or "vercel" in request.headers.get("Host", "").lower() else "http"
        host = request.headers.get("Host", "localhost:8000")
        portal_link = f"{scheme}://{host}/"
        
        emp_body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; padding: 20px; color: #1e293b; text-align: center;">
                <div style="background-color: white; padding: 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">
                    <h2 style="color: #ef4444; border-bottom: 2px solid #fef2f2; padding-bottom: 10px; margin-top: 0;">Offboarding Initiated</h2>
                    <p style="font-size: 15px; color: #475569; line-height: 1.6; text-align: left;">
                        Dear {emp['name']},<br><br>
                        Following the registration of your resignation, we have initialized your offboarding process.<br><br>
                        Please click the link below to access your personal Employee Portal dashboard and submit your custom notice period details:
                    </p>
                    <div style="margin: 30px 0;">
                        <a href="{portal_link}" style="background-color: #ef4444; color: white; padding: 12px 24px; text-decoration: none; border-radius: 8px; font-weight: bold;">Access Employee Portal</a>
                    </div>
                    <p style="font-size: 13px; color: #94a3b8; text-align: left; border-top: 1px solid #f1f5f9; padding-top: 15px; margin-top: 20px;">
                        Thank you for your service at SEMCO Groups.
                    </p>
                </div>
            </body>
        </html>
        """
        if target_email:
            send_email(target_email, emp_subject, emp_body)
            
        return jsonify({"message": f"Offboarding triggered successfully for {emp['name']}."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/offboarding/notice/<employee_id>', methods=['POST'])
@login_required
def submit_notice_period(employee_id):
    try:
        emp = db.employees.find_one({"_id": ObjectId(employee_id)})
        if not emp:
            return jsonify({"detail": "Employee not found"}), 404
            
        data = request.json or {}
        serve_notice = data.get("serve_notice", True)
        duration = data.get("duration", 0)
        unit = data.get("unit", "days")
        
        db.employees.update_one({"_id": ObjectId(employee_id)}, {"$set": {
            "status": "Notice Period Submitted",
            "notice_period_served": serve_notice,
            "notice_period_duration": duration,
            "notice_period_unit": unit
        }})
        
        return jsonify({"message": "Notice period details submitted successfully."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/offboarding/schedule-interview/<employee_id>', methods=['POST'])
@login_required
def schedule_exit_interview(employee_id):
    try:
        emp = db.employees.find_one({"_id": ObjectId(employee_id)})
        if not emp:
            return jsonify({"detail": "Employee not found"}), 404
            
        data = request.json or {}
        date = data.get("date")
        time = data.get("time")
        location = data.get("location", "Google Meet")
        
        if not date or not time:
            return jsonify({"detail": "Date and Time are required"}), 400
            
        db.employees.update_one({"_id": ObjectId(employee_id)}, {"$set": {
            "status": "Interview Scheduled",
            "exit_interview_date": date,
            "exit_interview_time": time,
            "exit_interview_location": location
        }})
        
        # Format date-time for template
        interview_datetime = f"{date} at {time}"
        first_name = emp['name'].split(' ')[0]
        
        # Exact requested email template
        emp_subject = "Offboarding Process & Checklist"
        emp_body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; padding: 20px; color: #1e293b;">
                <div style="background-color: white; padding: 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">
                    <h2 style="color: #0f172a; border-bottom: 2px solid #eff6ff; padding-bottom: 10px; margin-top: 0; text-align: center;">Employee Offboarding Checklist</h2>
                    <p style="font-size: 15px; color: #334155; line-height: 1.6; text-align: center; margin-top: 20px;">
                        Dear {first_name},<br><br>
                        Following the registration of your resignation, we have initialized your offboarding. Your exit interview has been scheduled for {interview_datetime}.
                    </p>
                    <h4 style="color:#0f172a; margin-top:25px; margin-bottom:10px; font-weight: bold;">Your Next Steps Checklist:</h4>
                    <ul style="font-size:14px; line-height:1.8; color:#334155; padding-left: 20px;">
                        <li style="margin-bottom: 8px;">Schedule handoff sessions with your direct team lead.</li>
                        <li style="margin-bottom: 8px;">Return all company assets (laptops, peripherals, badges) on or before your last day.</li>
                        <li style="margin-bottom: 8px;">Submit your final business expense claims.</li>
                        <li style="margin-bottom: 8px;">Review exit interview documentation.</li>
                    </ul>
                    <p style="font-size: 14px; color: #64748b; margin-top:30px; text-align: center; font-weight: bold;">
                        Thank you for your service at SEMCO Groups.
                    </p>
                </div>
            </body>
        </html>
        """
        
        target_email = emp.get("personal_email") or emp.get("email")
        if target_email:
            send_email(target_email, emp_subject, emp_body)
            
        return jsonify({"message": "Exit interview scheduled and notification email sent."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/offboarding/mark-checklist/<employee_id>', methods=['POST'])
@login_required
def mark_checklist_complete(employee_id):
    try:
        emp = db.employees.find_one({"_id": ObjectId(employee_id)})
        if not emp:
            return jsonify({"detail": "Employee not found"}), 404
            
        db.employees.update_one({"_id": ObjectId(employee_id)}, {"$set": {
            "status": "Checklist Completed"
        }})
        
        return jsonify({"message": "Exit interview and checklists marked as completed."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/offboarding/delete/<employee_id>', methods=['POST'])
@login_required
def delete_employee_offboarding(employee_id):
    try:
        emp = db.employees.find_one({"_id": ObjectId(employee_id)})
        if not emp:
            return jsonify({"detail": "Employee not found"}), 404
            
        # 1. Archive historical details into db.archive_employees
        archive_doc = dict(emp)
        archive_doc["status"] = "ARCHIVED"
        archive_doc["system_access_revoked"] = 1
        archive_doc["archived_at"] = datetime.datetime.utcnow().isoformat()
        
        db.archive_employees.insert_one(archive_doc)
        
        # 2. Revoke and delete from active employees collection
        db.employees.delete_one({"_id": ObjectId(employee_id)})
        
        # 3. Clear the invitation record so the same personal email can be
        #    re-used for a fresh onboarding invite in the future.
        personal_email_to_clear = emp.get("personal_email") or emp.get("email", "")
        if personal_email_to_clear:
            db.invitations.delete_many({"personal_email": personal_email_to_clear})
        
        return jsonify({"message": f"Employee {emp['name']} has been safely archived, deleted, and portal access revoked."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Pulse Surveys ---
@app.route('/api/surveys/metrics', methods=['GET'])
@login_required
def get_survey_metrics():
    try:
        tenant_id = g.current_user["tenant_id"]
        
        pipeline = [
            {"$match": {"tenant_id": tenant_id}},
            {"$group": {
                "_id": None,
                "b_avg": {"$avg": "$q1_burnout"},
                "a_avg": {"$avg": "$q2_alignment"},
                "s_avg": {"$avg": "$q3_satisfaction"},
                "cnt": {"$sum": 1}
            }}
        ]
        
        res = list(db.survey_responses.aggregate(pipeline))
        
        if not res:
            return jsonify({"burnout_avg": 0, "alignment_avg": 0, "satisfaction_avg": 0, "count": 0})
            
        data = res[0]
        return jsonify({
            "burnout_avg": round(data["b_avg"] or 0, 1),
            "alignment_avg": round(data["a_avg"] or 0, 1),
            "satisfaction_avg": round(data["s_avg"] or 0, 1),
            "count": data["cnt"]
        })
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/surveys/submit', methods=['POST'])
@login_required
def submit_survey():
    data = request.json or {}
    q1 = int(data.get('q1_burnout', 3))
    q2 = int(data.get('q2_alignment', 3))
    q3 = int(data.get('q3_satisfaction', 3))
    
    tenant_id = g.current_user["tenant_id"]
    now = datetime.datetime.now().isoformat()
    survey_month = datetime.datetime.now().strftime('%Y-%m')
    
    try:
        db.survey_responses.insert_one({
            "tenant_id": tenant_id,
            "survey_month": survey_month,
            "q1_burnout": q1,
            "q2_alignment": q2,
            "q3_satisfaction": q3,
            "submitted_at": now
        })
        return jsonify({"message": "Survey submitted successfully! Thank you for your feedback."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Centralized Policies (LLM KB) ---
@app.route('/api/policies', methods=['GET', 'POST'])
@login_required
def handle_policies():
    try:
        tenant_id = g.current_user["tenant_id"]
        if request.method == 'POST':
            category = request.form.get('category')
            title = request.form.get('title')
            content = request.form.get('content') or ''
            
            if not title or not category:
                return jsonify({"detail": "category and title are required"}), 400
                
            file_url = None
            file_name = None
            if 'file' in request.files:
                file = request.files['file']
                if file and file.filename != '':
                    from werkzeug.utils import secure_filename
                    import os
                    import uuid
                    filename = secure_filename(file.filename)
                    upload_dir = get_upload_dir('policies')
                    unique_filename = f"{uuid.uuid4().hex}_{filename}"
                    filepath = os.path.join(upload_dir, unique_filename)
                    file.save(filepath)
                    file_url = f"/static/uploads/policies/{unique_filename}"
                    file_name = filename
                    
            res = db.policies.insert_one({
                "tenant_id": tenant_id,
                "category": category,
                "title": title,
                "content": content,
                "file_url": file_url,
                "file_name": file_name
            })
            p = db.policies.find_one({"_id": res.inserted_id})
            return jsonify(serialize_doc(p))
        else:
            policies = list(db.policies.find({"tenant_id": tenant_id}))
            return jsonify(serialize_list(policies))
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/policies/search', methods=['GET'])
@login_required
def search_policies():
    query = request.args.get('q', '').strip()
    if not query:
        return jsonify({"answer": "Please ask a question about company policies."})
        
    try:
        tenant_id = g.current_user["tenant_id"]
        policies = list(db.policies.find({"tenant_id": tenant_id}))
        
        matches = []
        for p in policies:
            if query.lower() in p["title"].lower() or query.lower() in p["content"].lower() or p["category"].lower() in query.lower():
                matches.append(p)
                
        if not matches:
            return jsonify({"answer": "I couldn't find any specific company SOP matching your question. Please refer to physical documentation or consult People Operations."})
            
        top_match = matches[0]
        answer = f"According to the **{top_match['title']}** ({top_match['category']}):\n\n"
        answer += f"> \"{top_match['content']}\"\n\n"
        answer += "Is there anything specific you would like to clarify about this policy?"
        
        return jsonify({"answer": answer})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Interview Scheduler Assistant ---
@app.route('/api/interviews', methods=['GET', 'POST'])
@login_required
def handle_interviews():
    try:
        tenant_id = g.current_user["tenant_id"]
        if request.method == 'POST':
            data = request.json or {}
            candidate_name = data.get('candidate_name')
            candidate_email = data.get('candidate_email')
            interview_time_str = data.get('interview_time')
            jd_title = data.get('jd_title')
            
            if not candidate_name or not candidate_email or not interview_time_str or not jd_title:
                return jsonify({"detail": "All fields are required"}), 400
                
            meeting_link = f"https://meet.google.com/abc-{candidate_name.replace(' ', '-').lower()[:4]}-xyz"
            
            res = db.interviews.insert_one({
                "tenant_id": tenant_id,
                "candidate_name": candidate_name,
                "candidate_email": candidate_email,
                "interview_time": interview_time_str,
                "jd_title": jd_title,
                "meeting_link": meeting_link,
                "status": "SCHEDULED"
            })
            i_id = res.inserted_id
            
            # Email Candidate
            subject = f"Interview Invitation: {jd_title} - Acme Corp"
            body = f"""
            <html>
                <body style="font-family: Arial, sans-serif; background-color: #f8fafc; padding: 20px; color: #1e293b;">
                    <div style="background-color: white; padding: 30px; border-radius: 12px; max-width: 600px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">
                        <h2 style="color: #a855f7; border-bottom: 2px solid #f3e8ff; padding-bottom: 10px; margin-top: 0;">Interview Confirmation</h2>
                        <p style="font-size: 15px; color: #475569;">
                            Hi {candidate_name},<br><br>
                            We are excited to discuss the <b>{jd_title}</b> opportunity with you. Your interview has been scheduled:
                        </p>
                        <table style="width:100%; border-collapse:collapse; margin:20px 0; font-size:14px;">
                            <tr><td style="padding:6px; color:#64748b;"><b>Time:</b></td><td>{interview_time_str}</td></tr>
                            <tr><td style="padding:6px; color:#64748b;"><b>Meeting Link:</b></td><td><a href="{meeting_link}">{meeting_link}</a></td></tr>
                            <tr><td style="padding:6px; color:#64748b;"><b>Job Role:</b></td><td>{jd_title}</td></tr>
                        </table>
                        <p style="font-size: 13px; color: #94a3b8;">Please ensure your camera and microphone are working before joining the meeting.</p>
                    </div>
                </body>
            </html>
            """
            jd_file = f"temp_jd_{str(i_id)}.txt"
            with open(jd_file, "w") as f:
                f.write(f"JOB DESCRIPTION: {jd_title}\nCompany: Acme Corporation\nDescription: We are looking for an exceptional talent to join our team...")
                
            send_email(candidate_email, subject, body, attachment_path=jd_file, attachment_name=f"Job_Description_{jd_title.replace(' ', '_')}.txt")
            
            if os.path.exists(jd_file):
                try:
                    os.remove(jd_file)
                except OSError:
                    pass
                    
            return jsonify({
                "id": str(i_id),
                "candidate_name": candidate_name,
                "candidate_email": candidate_email,
                "interview_time": interview_time_str,
                "jd_title": jd_title,
                "meeting_link": meeting_link,
                "status": "SCHEDULED"
            })
        else:
            interviews = list(db.interviews.find({"tenant_id": tenant_id}))
            return jsonify(serialize_list(interviews))
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Attendance Endpoints ---
@app.route('/api/auth/signup-status', methods=['GET'])
def get_signup_status():
    try:
        admin_exists = db.employees.count_documents({"role": "Admin (HR)"}) > 0
        return jsonify({"admin_registered": admin_exists})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/attendance/today', methods=['GET'])
@login_required
def get_today_attendance():
    try:
        user_id = g.current_user["employee_id"]
        today_str = datetime.date.today().isoformat()
        
        record = db.attendance.find_one({"employee_id": ObjectId(user_id), "date": today_str})
        if not record:
            return jsonify({"date": today_str, "selections": []})
            
        selections = list(record.get("selections", {}).keys())
        return jsonify({"date": today_str, "selections": selections})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/attendance/mark', methods=['POST'])
@login_required
def mark_attendance():
    try:
        user_id = g.current_user["employee_id"]
        
        # Enforce lock status check
        today_date = datetime.date.today()
        lock_record = db.attendance_locks.find_one({"year": today_date.year, "month": f"{today_date.month:02d}"})
        if lock_record and lock_record.get("locked"):
            return jsonify({"detail": "This month's attendance has been locked and finalized by HR. No further selections can be submitted."}), 403

        data = request.json or {}
        selection = data.get("selection")
        
        valid_selections = [
            'Present',
            'Weekly Off',
            'Sick Leave',
            'Casual Leave',
            'Privileged Leave',
            'Site Visit',
            'Back From Site Visit',
            'Extended Work'
        ]
        
        if selection not in valid_selections:
            return jsonify({"detail": "Invalid status selection"}), 400
            
        now = datetime.datetime.now()
        current_time_str = now.strftime("%H:%M")
        is_in_window = ("10:00" <= current_time_str <= "10:30")
        
        if not is_in_window:
            emp = db.employees.find_one({"_id": ObjectId(user_id)})
            if not emp or not emp.get("allow_late_attendance_marking"):
                return jsonify({"detail": "Attendance portal is locked. Attendance can only be marked between 10:00 AM and 10:30 AM, unless authorized by an HR Admin."}), 403
            
        today_str = datetime.date.today().isoformat()
        time_str = now.strftime("%I:%M %p") # Time-wise AM/PM format
        
        # Check if record exists for today
        record = db.attendance.find_one({"employee_id": ObjectId(user_id), "date": today_str})
        
        if record:
            selections = record.get("selections", {})
            if selection in selections:
                return jsonify({"detail": f"Selection '{selection}' is already locked for today."}), 400
                
            # Lock-in rule: selections cannot be removed/deselected. We only add new checked items!
            selections[selection] = time_str
            db.attendance.update_one(
                {"_id": record["_id"]},
                {"$set": {"selections": selections}}
            )
        else:
            selections = {selection: time_str}
            db.attendance.insert_one({
                "employee_id": ObjectId(user_id),
                "date": today_str,
                "selections": selections
            })
            
        # Clear the late permit flag once successfully marked
        if not is_in_window:
            db.employees.update_one({"_id": ObjectId(user_id)}, {"$set": {"allow_late_attendance_marking": False}})
            
        return jsonify({"date": today_str, "selections": list(selections.keys())})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/attendance/permit/<employee_id>', methods=['POST'])
@login_required
def permit_late_attendance(employee_id):
    try:
        if g.current_user["role"] != "Admin (HR)":
            return jsonify({"detail": "Admin access required"}), 403
            
        emp = db.employees.find_one({"_id": ObjectId(employee_id)})
        if not emp:
            return jsonify({"detail": "Employee not found"}), 404
            
        db.employees.update_one(
            {"_id": ObjectId(employee_id)},
            {"$set": {"allow_late_attendance_marking": True}}
        )
        return jsonify({"message": f"Late attendance marking has been permitted for {emp['name']}."})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/attendance/my-month', methods=['GET'])
@login_required
def get_my_month_attendance():
    try:
        user_id = g.current_user["employee_id"]
        year = request.args.get("year", datetime.date.today().year)
        month = request.args.get("month", f"{datetime.date.today().month:02d}")
        
        month_prefix = f"{year}-{month}"
        
        records = list(db.attendance.find({
            "employee_id": ObjectId(user_id),
            "date": {"$regex": f"^{month_prefix}"}
        }))
        
        records_list = []
        for r in records:
            records_list.append({
                "date": r["date"],
                "selections": r.get("selections", {})
            })
            
        return jsonify({
            "records": records_list
        })
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/attendance/admin/month', methods=['GET'])
@login_required
def get_admin_month_attendance():
    try:
        if g.current_user["role"] != "Admin (HR)":
            return jsonify({"detail": "Admin access required"}), 403
            
        year = request.args.get("year", datetime.date.today().year)
        month = request.args.get("month", f"{datetime.date.today().month:02d}")
        
        month_prefix = f"{year}-{month}" # YYYY-MM
        
        # Find all attendance records matching month_prefix
        records_cursor = db.attendance.find({"date": {"$regex": f"^{month_prefix}"}})
        
        # Map employee names
        employees = list(db.employees.find({"tenant_id": "semco"}))
        emp_map = {str(e["_id"]): e["name"] for e in employees}
        
        records_list = []
        for r in records_cursor:
            emp_id_str = str(r["employee_id"])
            records_list.append({
                "employee_id": emp_id_str,
                "employee_name": emp_map.get(emp_id_str, "Unknown Employee"),
                "date": r["date"],
                "selections": r.get("selections", {})
            })
            
        return jsonify({
            "records": records_list,
            "employees": [{"id": str(e["_id"]), "name": e["name"], "allow_late_attendance_marking": e.get("allow_late_attendance_marking", False)} for e in employees]
        })
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/attendance/admin/export', methods=['GET'])
@login_required
def export_attendance():
    try:
        if g.current_user["role"] != "Admin (HR)":
            return jsonify({"detail": "Admin access required"}), 403
            
        year = request.args.get("year", datetime.date.today().year)
        month = request.args.get("month", f"{datetime.date.today().month:02d}")
        fmt = request.args.get("format", "csv").lower()
        
        from app.utils.attendance_export import generate_attendance_data, export_csv, export_xls, export_word, export_pdf
        
        summary_data, detailed_data = generate_attendance_data(db, year, month)
        
        if fmt == "csv":
            file_data = export_csv(summary_data, detailed_data)
            mimetype = "text/csv"
            filename = f"attendance_masterlist_{year}_{month}.csv"
        elif fmt == "xls":
            file_data = export_xls(summary_data, detailed_data)
            mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"attendance_masterlist_{year}_{month}.xlsx"
        elif fmt == "word":
            file_data = export_word(summary_data, detailed_data, year, month)
            mimetype = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            filename = f"attendance_masterlist_{year}_{month}.docx"
        elif fmt == "pdf":
            file_data = export_pdf(summary_data, detailed_data, year, month)
            mimetype = "application/pdf"
            filename = f"attendance_masterlist_{year}_{month}.pdf"
        else:
            return jsonify({"detail": "Invalid format requested"}), 400
            
        return send_file(
            io.BytesIO(file_data),
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Customized Monthly Leave Allocation, Tracking, and Payroll Sync ---
@app.route('/api/attendance/leave-allocation', methods=['POST'])
@login_required
def save_leave_allocation():
    try:
        if g.current_user["role"] != "Admin (HR)":
            return jsonify({"detail": "Admin access required"}), 403
            
        data = request.json or {}
        emp_id = data.get("employee_id") # "all" or specific employee ID
        year = int(data.get("year", datetime.date.today().year))
        month = data.get("month", f"{datetime.date.today().month:02d}")
        
        # Allocations
        wo = float(data.get("wo", 0))
        sl = float(data.get("sl", 0))
        cl = float(data.get("cl", 0))
        pl = float(data.get("pl", 0))
        
        if emp_id == "all":
            # Allocate to all employees
            employees = list(db.employees.find({"tenant_id": "semco"}))
            for emp in employees:
                db.leave_allocations.update_one(
                    {
                        "employee_id": emp["_id"],
                        "year": year,
                        "month": month
                    },
                    {
                        "$set": {
                            "wo": wo,
                            "sl": sl,
                            "cl": cl,
                            "pl": pl,
                            "updated_at": datetime.datetime.utcnow().isoformat()
                        }
                    },
                    upsert=True
                )
        else:
            db.leave_allocations.update_one(
                {
                    "employee_id": ObjectId(emp_id),
                    "year": year,
                    "month": month
                },
                {
                    "$set": {
                        "wo": wo,
                        "sl": sl,
                        "cl": cl,
                        "pl": pl,
                        "updated_at": datetime.datetime.utcnow().isoformat()
                    }
                },
                upsert=True
            )
            
        return jsonify({"message": "Leave allocations saved successfully!"})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/attendance/leave-allocation', methods=['GET'])
@login_required
def get_leave_allocations():
    try:
        year = int(request.args.get("year", datetime.date.today().year))
        month = request.args.get("month", f"{datetime.date.today().month:02d}")
        emp_id = request.args.get("employee_id")
        
        if emp_id:
            alloc = db.leave_allocations.find_one({
                "employee_id": ObjectId(emp_id),
                "year": year,
                "month": month
            })
            if not alloc:
                return jsonify({"wo": 0, "sl": 0, "cl": 0, "pl": 0})
            return jsonify({
                "wo": alloc.get("wo", 0),
                "sl": alloc.get("sl", 0),
                "cl": alloc.get("cl", 0),
                "pl": alloc.get("pl", 0)
            })
        else:
            allocs = list(db.leave_allocations.find({
                "year": year,
                "month": month
            }))
            res = {}
            for a in allocs:
                res[str(a["employee_id"])] = {
                    "wo": a.get("wo", 0),
                    "sl": a.get("sl", 0),
                    "cl": a.get("cl", 0),
                    "pl": a.get("pl", 0)
                }
            return jsonify(res)
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/attendance/leave-summary', methods=['GET'])
@login_required
def get_leave_summary():
    try:
        user_id = request.args.get("employee_id") or g.current_user["employee_id"]
        year = int(request.args.get("year", datetime.date.today().year))
        month = request.args.get("month", f"{datetime.date.today().month:02d}")
        
        alloc = db.leave_allocations.find_one({
            "employee_id": ObjectId(user_id),
            "year": year,
            "month": month
        })
        if not alloc:
            alloc = {"wo": 0, "sl": 0, "cl": 0, "pl": 0}
            
        month_prefix = f"{year}-{month}"
        records = list(db.attendance.find({
            "employee_id": ObjectId(user_id),
            "date": {"$regex": f"^{month_prefix}"}
        }))
        
        wo_consumed = 0
        sl_consumed = 0
        cl_consumed = 0
        pl_consumed = 0
        
        for r in records:
            sels = r.get("selections", {})
            if "Weekly Off" in sels:
                wo_consumed += 1
            if "Sick Leave" in sels:
                sl_consumed += 1
            if "Casual Leave" in sels:
                cl_consumed += 1
            if "Privileged Leave" in sels:
                pl_consumed += 1
                
        categories = [
            {
                "category": "Week Off",
                "allocated": alloc.get("wo", 0),
                "consumed": wo_consumed,
                "balance": max(0.0, float(alloc.get("wo", 0)) - wo_consumed)
            },
            {
                "category": "Sick Leave",
                "allocated": alloc.get("sl", 0),
                "consumed": sl_consumed,
                "balance": max(0.0, float(alloc.get("sl", 0)) - sl_consumed)
            },
            {
                "category": "Casual Leave",
                "allocated": alloc.get("cl", 0),
                "consumed": cl_consumed,
                "balance": max(0.0, float(alloc.get("cl", 0)) - cl_consumed)
            },
            {
                "category": "Privileged Leave",
                "allocated": alloc.get("pl", 0),
                "consumed": pl_consumed,
                "balance": max(0.0, float(alloc.get("pl", 0)) - pl_consumed)
            }
        ]
        
        return jsonify(categories)
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/attendance/lock', methods=['POST'])
@login_required
def lock_attendance_month():
    try:
        if g.current_user["role"] != "Admin (HR)":
            return jsonify({"detail": "Admin access required"}), 403
            
        data = request.json or {}
        year = int(data.get("year", datetime.date.today().year))
        month = data.get("month", f"{datetime.date.today().month:02d}")
        
        db.attendance_locks.update_one(
            {
                "year": year,
                "month": month
            },
            {
                "$set": {
                    "locked": True,
                    "locked_at": datetime.datetime.utcnow().isoformat(),
                    "locked_by": ObjectId(g.current_user["employee_id"])
                }
            },
            upsert=True
        )
        
        employees = list(db.employees.find({"tenant_id": "semco", "status": "ACTIVE"}))
        month_prefix = f"{year}-{month}"
        pay_period = f"{year}-{month}"
        
        synced_count = 0
        for emp in employees:
            emp_id = emp["_id"]
            
            records = list(db.attendance.find({
                "employee_id": emp_id,
                "date": {"$regex": f"^{month_prefix}"}
            }))
            
            present_count = 0
            wo_consumed = 0
            sl_consumed = 0
            cl_consumed = 0
            pl_consumed = 0
            
            for r in records:
                sels = r.get("selections", {})
                if any(x in sels for x in ["Present", "Site Visit", "Back From Site Visit", "Extended Work"]):
                    present_count += 1
                if "Weekly Off" in sels:
                    wo_consumed += 1
                if "Sick Leave" in sels:
                    sl_consumed += 1
                if "Casual Leave" in sels:
                    cl_consumed += 1
                if "Privileged Leave" in sels:
                    pl_consumed += 1
            
            alloc = db.leave_allocations.find_one({
                "employee_id": emp_id,
                "year": year,
                "month": month
            })
            if not alloc:
                alloc = {"wo": 0, "sl": 0, "cl": 0, "pl": 0}
                
            wo_alloc = float(alloc.get("wo", 0))
            sl_alloc = float(alloc.get("sl", 0))
            cl_alloc = float(alloc.get("cl", 0))
            pl_alloc = float(alloc.get("pl", 0))
            
            lwp_days = max(0.0, sl_consumed - sl_alloc) + max(0.0, cl_consumed - cl_alloc) + max(0.0, pl_consumed - pl_alloc)
            total_leaves_taken = sl_consumed + cl_consumed + pl_consumed
            total_leaves_alloc = sl_alloc + cl_alloc + pl_alloc
            leaves_balance = max(0.0, total_leaves_alloc - total_leaves_taken)
            
            existing_payroll = db.payrolls.find_one({"employee_id": emp_id, "pay_period": pay_period})
            
            base_salary = 25000.0
            hra = 10000.0
            special_allowance = 5000.0
            conveyance_allowance = 1200.0
            other_allowance = 0.0
            reimbursment = 0.0
            advance_deducted = 0.0
            mlwf = 0.0
            pf = 1800.0
            esi = 0.0
            pt = 200.0
            
            if existing_payroll:
                base_salary = existing_payroll.get("base_salary", base_salary)
                hra = existing_payroll.get("hra", hra)
                special_allowance = existing_payroll.get("special_allowance", special_allowance)
                conveyance_allowance = existing_payroll.get("conveyance_allowance", conveyance_allowance)
                other_allowance = existing_payroll.get("other_allowance", other_allowance)
                reimbursment = existing_payroll.get("reimbursment", reimbursment)
                advance_deducted = existing_payroll.get("advance_decucted", advance_deducted)
                mlwf = existing_payroll.get("mlwf", mlwf)
                pf = existing_payroll.get("pf", pf)
                esi = existing_payroll.get("esi", esi)
                pt = existing_payroll.get("pt", pt)
            else:
                prev_payroll = db.payrolls.find_one({"employee_id": emp_id}, sort=[("pay_period", -1)])
                if prev_payroll:
                    base_salary = prev_payroll.get("base_salary", base_salary)
                    hra = prev_payroll.get("hra", hra)
                    special_allowance = prev_payroll.get("special_allowance", special_allowance)
                    conveyance_allowance = prev_payroll.get("conveyance_allowance", conveyance_allowance)
                    other_allowance = prev_payroll.get("other_allowance", other_allowance)
                    reimbursment = prev_payroll.get("reimbursment", reimbursment)
                    advance_deducted = prev_payroll.get("advance_decucted", advance_deducted)
                    mlwf = prev_payroll.get("mlwf", mlwf)
                    pf = prev_payroll.get("pf", pf)
                    esi = prev_payroll.get("esi", esi)
                    pt = prev_payroll.get("pt", pt)
            
            lwp_deduction = round(lwp_days * (base_salary / 30.0), 2)
            
            gross = base_salary + hra + special_allowance + other_allowance + conveyance_allowance + reimbursment
            total_deductions = advance_deducted + mlwf + pf + esi + pt + lwp_deduction
            net_salary = gross - total_deductions
            
            payload = {
                "present_days": present_count,
                "leaves_taken": total_leaves_taken,
                "leaves_balance": leaves_balance,
                "lwp_days": lwp_days,
                "lwp_deduction": lwp_deduction,
                "base_salary": base_salary,
                "basic_salary": base_salary,
                "hra": hra,
                "special_allowance": special_allowance,
                "other_allowance": other_allowance,
                "conveyance_allowance": conveyance_allowance,
                "reimbursment": reimbursment,
                "advance_decucted": advance_deducted,
                "mlwf": mlwf,
                "pf": pf,
                "esi": esi,
                "pt": pt,
                "allowances": gross - base_salary,
                "deductions": total_deductions,
                "net_salary": net_salary,
                "pay_period": pay_period,
                "status": "PENDING"
            }
            
            if existing_payroll:
                db.payrolls.update_one({"_id": existing_payroll["_id"]}, {"$set": payload})
            else:
                payload["employee_id"] = emp_id
                db.payrolls.insert_one(payload)
                
            synced_count += 1
            
        return jsonify({
            "message": f"Attendance locked successfully and synced to Payroll Engine for {synced_count} active employees!"
        })
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

@app.route('/api/attendance/lock-status', methods=['GET'])
@login_required
def get_attendance_lock_status():
    try:
        year = int(request.args.get("year", datetime.date.today().year))
        month = request.args.get("month", f"{datetime.date.today().month:02d}")
        
        lock = db.attendance_locks.find_one({
            "year": year,
            "month": month
        })
        if lock:
            return jsonify({"locked": lock.get("locked", False), "locked_at": lock.get("locked_at")})
        return jsonify({"locked": False})
    except Exception as e:
        return jsonify({"detail": str(e)}), 500

# --- Onboarding & Invitation API Routes ---
import uuid

@app.route('/api/invite/send', methods=['POST'])
@login_required
def send_onboarding_invite():
    try:
        if g.current_user["role"] != "Admin (HR)":
            return jsonify({"detail": "Admin access required"}), 403
            
        data = request.json or {}
        salutation = data.get('salutation', 'Mr.')
        name = data.get('name', '')
        personal_email = data.get('personal_email', '')
        
        if not name or not personal_email:
            return jsonify({"detail": "Name and Personal Email are required."}), 400
            
        # Check if email is already invited or in roster
        existing_invite = db.invitations.find_one({"personal_email": personal_email})
        
        if existing_invite and existing_invite.get("status") != "SENT":
            # Allow re-onboarding if this person was previously offboarded
            # (they will no longer have an active employee record)
            still_active = db.employees.find_one(
                {"$or": [{"personal_email": personal_email}, {"email": personal_email}]}
            )
            if still_active:
                return jsonify({"detail": "An onboarding profile has already been created for this email and the employee is still active."}), 400
            # Employee was offboarded — purge the stale invite so a fresh one can be sent
            db.invitations.delete_many({"personal_email": personal_email})
            existing_invite = None
            
        token = str(uuid.uuid4())
        
        db.invitations.update_one(
            {"personal_email": personal_email},
            {"$set": {
                "salutation": salutation,
                "name": name,
                "token": token,
                "status": "SENT",
                "created_at": datetime.datetime.utcnow().isoformat()
            }},
            upsert=True
        )
        
        scheme = "https" if request.is_secure or "vercel" in request.headers.get("Host", "").lower() else "http"
        host = request.headers.get("Host", "localhost:8000")
        invite_link = f"{scheme}://{host}/invite/activate?token={token}"
            
        subject = "Invitation to Onboard at SEMCO Groups"
        body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.8; color: #1a1a1a; background-color: #f9f9f9; margin: 0; padding: 0;">
                <div style="max-width: 620px; margin: 40px auto; background: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.08);">
                    <div style="background: linear-gradient(135deg, #15803d 0%, #1d4ed8 100%); padding: 36px 40px;">
                        <h2 style="color: #ffffff; margin: 0; font-size: 28px; font-weight: 800; letter-spacing: -0.5px;">Welcome to SEMCO Groups!</h2>
                        <p style="color: rgba(255,255,255,0.85); margin: 8px 0 0 0; font-size: 17px;">Your official employee onboarding invitation</p>
                    </div>
                    <div style="padding: 40px;">
                        <p style="font-size: 19px; margin: 0 0 20px 0; color: #1a1a1a;">Hello <strong>{salutation} {name}</strong>,</p>
                        <p style="font-size: 18px; color: #333333; margin: 0 0 20px 0; line-height: 1.8;">
                            We are thrilled to welcome you to our team. To begin your onboarding process, please click the secure button below to self-populate your profile details in our HR system:
                        </p>
                        <div style="text-align: center; margin: 35px 0;">
                            <a href="{invite_link}" style="display: inline-block; padding: 16px 36px; background: linear-gradient(135deg, #15803d, #1d4ed8); color: white; text-decoration: none; border-radius: 10px; font-weight: 800; font-size: 18px; letter-spacing: 0.3px; box-shadow: 0 6px 16px rgba(21, 128, 61, 0.3);">
                                Complete Onboarding Form
                            </a>
                        </div>
                        <p style="font-size: 16px; color: #555555; margin: 20px 0 8px 0;">If the button above does not work, copy and paste the link below into your browser:</p>
                        <p style="background: #f3f4f6; padding: 14px 16px; border-radius: 8px; font-family: monospace; font-size: 15px; color: #1d4ed8; word-break: break-all; border: 1px solid #e5e7eb;">{invite_link}</p>

                        <hr style="border: none; border-top: 2px solid #e5e7eb; margin: 32px 0;" />

                        <h3 style="color: #1d4ed8; margin: 0 0 12px 0; font-size: 18px;">Documents You Will Need to Submit</h3>
                        <p style="font-size: 15px; color: #475569; margin: 0 0 16px 0;">Once your profile is set up, you will be prompted to upload the following documents through the Employee Document Vault in the HR portal. Please have these ready in digital format (PDF / JPG / PNG):</p>

                        <table style="width: 100%; border-collapse: collapse; font-size: 14px; color: #1e293b;">
                          <thead>
                            <tr style="background: #f1f5f9;">
                              <th style="text-align: left; padding: 10px 14px; border-bottom: 2px solid #e2e8f0;">Document</th>
                              <th style="text-align: left; padding: 10px 14px; border-bottom: 2px solid #e2e8f0;">Format</th>
                              <th style="text-align: left; padding: 10px 14px; border-bottom: 2px solid #e2e8f0;">Required?</th>
                            </tr>
                          </thead>
                          <tbody>
                            <tr style="border-bottom: 1px solid #f1f5f9;">
                              <td style="padding: 10px 14px;">Aadhaar Card</td>
                              <td style="padding: 10px 14px;">PDF</td>
                              <td style="padding: 10px 14px; color: #ef4444; font-weight: 600;">Mandatory</td>
                            </tr>
                            <tr style="background: #f8fafc; border-bottom: 1px solid #f1f5f9;">
                              <td style="padding: 10px 14px;">PAN Card</td>
                              <td style="padding: 10px 14px;">PDF</td>
                              <td style="padding: 10px 14px; color: #ef4444; font-weight: 600;">Mandatory</td>
                            </tr>
                            <tr style="border-bottom: 1px solid #f1f5f9;">
                              <td style="padding: 10px 14px;">Profile Photo</td>
                              <td style="padding: 10px 14px;">JPG / PNG</td>
                              <td style="padding: 10px 14px; color: #ef4444; font-weight: 600;">Mandatory</td>
                            </tr>
                            <tr style="background: #f8fafc; border-bottom: 1px solid #f1f5f9;">
                              <td style="padding: 10px 14px;">Academic Proofs (10th, 12th, Degree certificates)</td>
                              <td style="padding: 10px 14px;">PDF / JPG / PNG</td>
                              <td style="padding: 10px 14px; color: #ef4444; font-weight: 600;">Mandatory</td>
                            </tr>
                            <tr style="border-bottom: 1px solid #f1f5f9;">
                              <td style="padding: 10px 14px;">Bank Passbook / Cancelled Cheque</td>
                              <td style="padding: 10px 14px;">PDF / JPG / PNG</td>
                              <td style="padding: 10px 14px; color: #64748b;">Optional</td>
                            </tr>
                            <tr style="background: #f8fafc;">
                              <td style="padding: 10px 14px;">Experience / Relieving Letter (if applicable)</td>
                              <td style="padding: 10px 14px;">PDF</td>
                              <td style="padding: 10px 14px; color: #64748b;">Optional</td>
                            </tr>
                          </tbody>
                        </table>

                        <div style="background: #eff6ff; border-left: 4px solid #1d4ed8; padding: 14px 18px; border-radius: 6px; margin: 24px 0; font-size: 14px; color: #1e40af;">
                          <strong>Bank Details Required:</strong> You will also be asked to provide your bank account number, account name, bank name, and IFSC code for payroll processing.
                        </div>

                        <hr style="border: none; border-top: 1px solid #e5e7eb; margin: 30px 0;" />
                        <p style="font-size: 17px; color: #333333; margin: 0 0 6px 0;">Best regards,</p>
                        <p style="font-size: 18px; color: #1a1a1a; font-weight: 700; margin: 0;">HR Operations Team<br/><span style="color: #15803d;">SEMCO Groups</span></p>
                    </div>
                    <footer style="background: #f8fafc; padding: 18px 40px; text-align: center; color: #94a3b8; font-size: 11px; border-top: 1px solid #e5e7eb;">
                      This invitation is private and non-transferable. Please do not share your onboarding link. &copy; SEMCO Groups HR Operations.
                    </footer>
                </div>
            </body>
        </html>
        """
        
        send_email(personal_email, subject, body)
        return jsonify({"message": "Invitation sent successfully."})
    except Exception as e:
        import traceback
        print(f"[ERROR] Onboarding invite failed:\n{traceback.format_exc()}")
        return jsonify({"detail": str(e)}), 500

@app.route('/api/invite/verify', methods=['GET'])
def verify_onboarding_token():
    try:
        token = request.args.get("token")
        if not token:
            return jsonify({"detail": "Token parameter is missing."}), 400
            
        invite = db.invitations.find_one({"token": token, "status": "SENT"})
        if not invite:
            return jsonify({"detail": "Invitation token is invalid, expired, or already used."}), 404

        return jsonify({
            "salutation": invite.get("salutation", "Mr."),
            "name": invite.get("name", ""),
            "personal_email": invite.get("personal_email", "")
        })
    except Exception as e:
        return jsonify({"detail": str(e)}), 500



@app.route('/api/invite/submit', methods=['POST'])
def submit_onboarding_info():

    try:
        # Accept both multipart/form-data (with files) and application/json (legacy)
        if request.content_type and 'multipart/form-data' in request.content_type:
            data = request.form
        else:
            data = request.json or {}

        token = data.get("token")
        if not token:
            return jsonify({"detail": "Invitation token is missing."}), 400

        invite = db.invitations.find_one({"token": token, "status": "SENT"})
        if not invite:
            return jsonify({"detail": "Invitation token is invalid, expired, or already used."}), 404

        salutation = data.get('salutation', 'Mr.')
        name = data.get('name', '')
        emp_id = data.get('emp_id', '')
        department = data.get('department', 'General')
        designation = data.get('designation', '')
        doj = data.get('doj', '')
        dob = data.get('dob', '')
        age = data.get('age', '')
        personal_email = data.get('personal_email', '')
        current_address = data.get('current_address', '')
        office_contact = data.get('office_contact', '')
        personal_contact = data.get('personal_contact', '')

        if dob and not age:
            try:
                dob_date = datetime.date.fromisoformat(str(dob)[:10])
                today = datetime.date.today()
                age = today.year - dob_date.year - ((today.month, today.day) < (dob_date.month, dob_date.day))
            except Exception:
                pass

        new_emp = {
            "salutation": salutation,
            "name": name,
            "emp_id": emp_id,
            "email": "",
            "password": "",
            "role": "Employee",
            "department": department,
            "designation": designation,
            "dob": dob,
            "doj": doj,
            "birthday": dob,
            "joining_date": doj,
            "age": age,
            "personal_email": personal_email,
            "current_address": current_address,
            "office_contact": office_contact,
            "personal_contact": personal_contact,
            "status": "PENDING",
            "system_access_revoked": 0,
            "tenant_id": "semco"
        }

        db.employees.update_one(
            {"personal_email": personal_email},
            {
                "$set": new_emp,
                "$setOnInsert": {"allow_personal_email_access": False}
            },
            upsert=True
        )

        # Fetch the newly created / updated employee _id so we can link the vault record
        created_emp = db.employees.find_one({"personal_email": personal_email})

        if created_emp:
            emp_oid = created_emp["_id"]

            # ── Process uploaded document files ────────────────────────────────
            from werkzeug.utils import secure_filename
            import uuid as _uuid

            upload_dir = get_upload_dir('documents')

            def save_onboard_file(field_key, allowed_ext=None):
                f = request.files.get(field_key)
                if not f or f.filename == '':
                    return None
                fname = f.filename.lower()
                if allowed_ext:
                    ext = os.path.splitext(fname)[1]
                    if ext not in allowed_ext:
                        return None
                b64_uri = convert_file_to_base64_uri(f, allowed_ext)
                return b64_uri

            # Collect personal document fields submitted during onboarding
            bank_name        = data.get('bank_name', '')
            account_number   = data.get('account_number', '')
            account_name     = data.get('account_name', '')
            ifsc_code        = data.get('ifsc_code', '')
            has_experience   = data.get('has_experience', 'false') == 'true'
            try:
                pct_10th  = float(data.get('percentage_10th', 0) or 0)
                pct_12th  = float(data.get('percentage_12th', 0) or 0)
                b_cgpa    = float(data.get('bachelors_cgpa', 0) or 0)
            except (ValueError, TypeError):
                pct_10th = pct_12th = b_cgpa = 0.0

            aadhaar_url         = save_onboard_file('aadhaar',         ['.pdf'])
            pan_url             = save_onboard_file('pan',             ['.pdf'])
            profile_photo_url   = save_onboard_file('profile_photo',   ['.jpg', '.jpeg', '.png'])
            academic_proofs_url = save_onboard_file('academic_proofs', ['.pdf', '.jpg', '.jpeg', '.png'])
            passbook_url        = save_onboard_file('passbook',        ['.pdf', '.jpg', '.jpeg', '.png'])
            experience_url      = save_onboard_file('experience',      ['.pdf'])

            # Build partial personal_documents dict (only fields that were actually provided)
            personal_data = {}
            if aadhaar_url:         personal_data['aadhaar_url']         = aadhaar_url
            if pan_url:             personal_data['pan_url']             = pan_url
            if profile_photo_url:   personal_data['profile_photo_url']   = profile_photo_url
            if academic_proofs_url: personal_data['academic_proofs_url'] = academic_proofs_url
            if passbook_url:        personal_data['passbook_url']        = passbook_url
            if experience_url:      personal_data['experience_url']      = experience_url
            if bank_name:           personal_data['bank_name']           = bank_name
            if account_number:      personal_data['account_number']      = account_number
            if account_name:        personal_data['account_name']        = account_name
            if ifsc_code:           personal_data['ifsc_code']           = ifsc_code
            if pct_10th:            personal_data['percentage_10th']     = pct_10th
            if pct_12th:            personal_data['percentage_12th']     = pct_12th
            if b_cgpa:              personal_data['bachelors_cgpa']      = b_cgpa
            personal_data['has_experience'] = has_experience

            # Upsert vault record — if it already exists (from auto-create), merge personal_documents
            existing_vault = db.employee_documents.find_one({"employee_id": emp_oid, "tenant_id": "semco"})
            if existing_vault:
                merged_personal = {**existing_vault.get("personal_documents", {}), **personal_data}
                db.employee_documents.update_one(
                    {"employee_id": emp_oid, "tenant_id": "semco"},
                    {"$set": {"personal_documents": merged_personal}}
                )
            else:
                db.employee_documents.insert_one({
                    "employee_id": emp_oid,
                    "tenant_id": "semco",
                    "personal_documents": personal_data,
                    "company_documents": [],
                    "created_at": datetime.datetime.utcnow().isoformat()
                })

        db.invitations.update_one({"token": token}, {"$set": {"status": "SUBMITTED"}})

        return jsonify({"message": "Onboarding information submitted successfully!"})
    except Exception as e:
        import traceback
        print(f"[ERROR] Onboarding submit failed:\n{traceback.format_exc()}")
        return jsonify({"detail": str(e)}), 500

# Catch-all route to serve Vite SPA frontend
# This handles ALL non-API routes, including SPA client routes like /invite/activate
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def catch_all(path):
    print(f"[DEBUG] catch_all path received: '{path}'")
    # Serve backend static uploads (certificates, documents, etc.)
    if path.startswith("static/"):
        if (os.environ.get("VERCEL") or os.environ.get("VERCEL_ENV")) and path.startswith("static/uploads/"):
            tmp_path = os.path.join("/tmp", path)
            if os.path.isfile(tmp_path):
                return send_from_directory("/tmp", path)
        requested_file = os.path.join(app.root_path, path)
        if os.path.isfile(requested_file):
            return send_from_directory(app.root_path, path)
        return jsonify({"detail": "Requested static document file not found."}), 404
            
    # Let API routes 404 cleanly
    if path.startswith("api/"):
        return jsonify({"detail": "Not found"}), 404
    # Try to serve a real static file from the dist folder (JS, CSS, images, etc.)
    requested_file = os.path.join(frontend_dist, path)
    if path and os.path.isfile(requested_file):
        return send_from_directory(frontend_dist, path)
    # For all SPA client-side routes (e.g. /invite/activate, /dashboard), serve index.html
    # Always send with no-cache so browsers pick up new builds immediately
    index_path = os.path.join(frontend_dist, "index.html")
    if os.path.isfile(index_path):
        resp = send_from_directory(frontend_dist, "index.html")
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp
    return jsonify({"detail": "Frontend not built. Run: cd frontend && npm run build"}), 404

if __name__ == '__main__':
    app.run(port=8000, debug=True)
